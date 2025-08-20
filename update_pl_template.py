#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
수정된 PL 추출 로직으로 자금수지표 템플릿 업데이트
- 첫 5개 PL 계정: 40000, 80200, 80800, 81100, 81200
- 연도: 2022, 2023, 2024
- 연말 손익대체 거래 자동 제외
"""

import sys
import os
sys.path.append('C:\\K04_cashflow_1\\Script')

from main_processor import LedgerExtractionEngine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
from datetime import datetime

def setup_logging():
    """UTF-8 로깅 설정"""
    log_filename = f"C:\\K04_cashflow_1\\log\\{datetime.now().strftime('%Y-%m-%d_%H%M%S')}_PL템플릿업데이트.log"
    os.makedirs("C:\\K04_cashflow_1\\log", exist_ok=True)
    
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='[%(asctime)s] [%(levelname)s] %(message)s',
        filemode='w',
        encoding='utf-8'
    )
    
    # 콘솔 출력도 추가
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('[%(levelname)s] %(message)s'))
    logging.getLogger().addHandler(console_handler)
    
    return log_filename

def update_pl_template():
    """PL 계정 데이터를 자금수지표 템플릿에 업데이트"""
    
    log_file = setup_logging()
    logging.info("[PL템플릿업데이트시작] [CLAUDE.md준수] [연말대체제외로직적용]")
    
    try:
        # 1. 추출 엔진 초기화
        extractor = LedgerExtractionEngine()
        logging.info("[원장추출엔진초기화] [연말대체감지기능포함]")
        
        # 2. 처리 대상 설정
        target_accounts = ['40000', '80200', '80800', '81100', '81200']
        target_years = [2022, 2023, 2024]
        template_path = "C:\\K04_cashflow_1\\Books\\자금수지표_v1_250820.xlsx"
        
        logging.info(f"[처리대상설정] [계정_{len(target_accounts)}개] [연도_{len(target_years)}개년]")
        logging.info(f"[대상계정] {target_accounts}")
        logging.info(f"[대상연도] {target_years}")
        
        # 3. 템플릿 파일 로드
        template_wb = load_workbook(template_path)
        logging.info(f"[템플릿로드완료] [파일_{template_path}]")
        
        # 4. PL 시트 확인/생성
        if 'PL계정' in template_wb.sheetnames:
            pl_sheet = template_wb['PL계정']
            logging.info("[기존PL시트사용]")
        else:
            pl_sheet = template_wb.create_sheet("PL계정")
            logging.info("[새PL시트생성]")
        
        # 5. 헤더 설정
        headers = ['계정코드', '계정명', '구분'] + [f'{year}년_전기이월' for year in target_years]
        for year in target_years:
            for month in range(1, 13):
                headers.append(f'{year}년_{month:02d}월')
        
        # 헤더 입력
        for col_idx, header in enumerate(headers, 1):
            pl_sheet.cell(row=1, column=col_idx).value = header
        
        logging.info(f"[헤더설정완료] [컬럼수_{len(headers)}개]")
        
        # 6. 각 계정별 데이터 추출 및 입력
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row_idx = 2
        
        for account_code in target_accounts:
            logging.info(f"[계정처리시작] [코드_{account_code}]")
            
            # 계정 분류 확인
            account_type = extractor.classify_account_type(account_code)
            
            if account_type == 'SUMMARY':
                logging.warning(f"[SUMMARY계정제외] [코드_{account_code}] [건너뛰기]")
                # SUMMARY 계정도 표시하되 처리 제외 표시
                pl_sheet.cell(row=row_idx, column=1).value = account_code
                pl_sheet.cell(row=row_idx, column=2).value = "손익(SUMMARY)"
                pl_sheet.cell(row=row_idx, column=3).value = "처리제외"
                # 노란색 마킹
                for col in range(4, len(headers) + 1):
                    pl_sheet.cell(row=row_idx, column=col).fill = yellow_fill
                row_idx += 1
                continue
            
            # 계정명 매핑
            account_names = {
                '80200': '직원급여',
                '80800': '퇴직급여',
                '81100': '복리후생비',
                '81200': '여비교통비'
            }
            
            # 기본 정보 입력
            pl_sheet.cell(row=row_idx, column=1).value = account_code
            pl_sheet.cell(row=row_idx, column=2).value = account_names.get(account_code, f'계정{account_code}')
            pl_sheet.cell(row=row_idx, column=3).value = account_type
            
            col_idx = 4  # 데이터 시작 컬럼
            
            # 7. 연도별 데이터 추출 및 입력
            for year in target_years:
                logging.info(f"[연도처리시작] [계정_{account_code}] [연도_{year}]")
                
                try:
                    # 원장 파일 로딩
                    ledger_file = extractor.account_mapping.get(year)
                    if not ledger_file or not os.path.exists(ledger_file):
                        logging.warning(f"[원장파일없음] [연도_{year}] [파일_{ledger_file}]")
                        # 전기이월 + 12개월을 노란색으로 마킹
                        for month_col in range(col_idx, col_idx + 13):
                            pl_sheet.cell(row=row_idx, column=month_col).fill = yellow_fill
                        col_idx += 13
                        continue
                    
                    workbook = load_workbook(ledger_file)
                    
                    # 계정 시트 찾기
                    account_sheet = None
                    for sheet_name in workbook.sheetnames:
                        if f'({account_code})' in sheet_name:
                            account_sheet = workbook[sheet_name]
                            break
                    
                    if not account_sheet:
                        logging.warning(f"[계정시트없음] [계정_{account_code}] [연도_{year}]")
                        # 전기이월 + 12개월을 노란색으로 마킹
                        for month_col in range(col_idx, col_idx + 13):
                            pl_sheet.cell(row=row_idx, column=month_col).fill = yellow_fill
                        col_idx += 13
                        continue
                    
                    # 전기이월 추출
                    carry_forward = extractor.extract_carry_forward(account_sheet)
                    
                    # 전기이월 입력
                    if carry_forward is not None:
                        pl_sheet.cell(row=row_idx, column=col_idx).value = carry_forward
                        logging.info(f"[전기이월입력] [계정_{account_code}] [연도_{year}] [금액_{carry_forward:,}원]")
                    else:
                        pl_sheet.cell(row=row_idx, column=col_idx).fill = yellow_fill
                        logging.warning(f"[전기이월불확실] [계정_{account_code}] [연도_{year}] [노란색마킹]")
                    
                    col_idx += 1
                    
                    # 월별 데이터 추출 (수정된 로직 적용)
                    monthly_data = extractor.extract_monthly_data(account_sheet, account_type, year=year)
                    
                    # 월별 데이터 입력
                    for month in range(1, 13):
                        month_amount = monthly_data.get(month, 0)
                        if month_amount != 0:
                            pl_sheet.cell(row=row_idx, column=col_idx).value = month_amount
                            if month == 12:
                                logging.info(f"[12월데이터확인] [계정_{account_code}] [연도_{year}] [금액_{month_amount:,}원]")
                        col_idx += 1
                    
                    workbook.close()
                    logging.info(f"[연도처리완료] [계정_{account_code}] [연도_{year}]")
                
                except Exception as e:
                    logging.error(f"[연도처리오류] [계정_{account_code}] [연도_{year}] [오류_{str(e)}]")
                    # 해당 연도 데이터를 노란색으로 마킹
                    for month_col in range(col_idx, col_idx + 13):
                        if month_col <= len(headers):
                            pl_sheet.cell(row=row_idx, column=month_col).fill = yellow_fill
                    col_idx += 13
            
            logging.info(f"[계정처리완료] [코드_{account_code}]")
            row_idx += 1
        
        # 8. 템플릿 파일 저장
        template_wb.save(template_path)
        template_wb.close()
        
        logging.info(f"[템플릿업데이트완료] [파일_{template_path}]")
        logging.info(f"[로그파일_{log_file}]")
        
        print("\n=== PL 템플릿 업데이트 완료 ===")
        print(f"대상 계정: {target_accounts}")
        print(f"대상 연도: {target_years}")
        print(f"템플릿 파일: {template_path}")
        print(f"로그 파일: {log_file}")
        print("\n✅ 연말 손익대체 거래 제외 로직 적용됨")
        print("✅ 12월 음수값 문제 해결됨")
        print("✅ CLAUDE.md 데이터 완결성 원칙 준수")
        
        return template_path
        
    except Exception as e:
        logging.error(f"[템플릿업데이트실패] [오류_{str(e)}]")
        print(f"템플릿 업데이트 실패: {e}")
        return None

if __name__ == "__main__":
    print("=== PL 계정 템플릿 업데이트 시작 ===")
    print("수정된 로직 적용: 연말 손익대체 거래 자동 제외")
    
    result = update_pl_template()
    
    if result:
        print(f"\n템플릿 업데이트 성공: {result}")
    else:
        print("\n템플릿 업데이트 실패")
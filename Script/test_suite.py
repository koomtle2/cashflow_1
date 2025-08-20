"""
재무데이터 처리 시스템 - 통합 테스트 스위트
CLAUDE.md 규칙 준수 검증 및 시스템 안정성 테스트
"""

import unittest
import tempfile
import os
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
import sys

# 테스트 대상 모듈들
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from logging_system import UTF8LoggingSystem
from validation_framework import DataValidator
from marking_system import YellowMarkingSystem
from mcp_interface import MCPInterface
from batch_processor import BatchProcessor
from main_processor import MainProcessor

class TestDataCreator:
    """테스트용 데이터 생성 클래스"""
    
    @staticmethod
    def create_sample_ledger_workbook() -> Workbook:
        """샘플 원장 워크북 생성"""
        wb = Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 테스트용 계정 시트들 생성
        test_accounts = [
            {'name': '보통예금(10300)', 'type': 'BS', 'has_carry_forward': True},
            {'name': '매출(41000)', 'type': 'PL', 'has_carry_forward': False},
            {'name': '부가세대급금(13500)', 'type': 'VAT', 'has_carry_forward': False},
            {'name': '잘못된계정(XXXXX)', 'type': 'UNKNOWN', 'has_carry_forward': False}
        ]
        
        for account in test_accounts:
            sheet = wb.create_sheet(account['name'])
            TestDataCreator._populate_account_sheet(
                sheet, account['type'], account['has_carry_forward']
            )
        
        return wb
    
    @staticmethod
    def _populate_account_sheet(sheet, account_type: str, has_carry_forward: bool):
        """계정 시트에 테스트 데이터 추가"""
        # 헤더
        headers = ['날짜', '적요', '거래처', '관리번호', '차변', '대변', '잔액']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col).value = header
        
        row = 2
        
        # 전기이월 (필요시)
        if has_carry_forward:
            sheet.cell(row=5, column=2).value = "전기이월"
            sheet.cell(row=5, column=7).value = 1000000  # 100만원
            row = 6
        
        # 월별 테스트 거래 데이터
        months = ['01', '02', '03', '04', '05', '06']
        balance = 1000000 if has_carry_forward else 0
        
        for month in months:
            # 월별로 2-3개 거래 생성
            for i in range(2):
                date_str = f"{month}-{i+1:02d}"
                debit = 50000 * (i + 1) if account_type != 'PL' else 0
                credit = 30000 * (i + 1) if account_type != 'PL' else 0
                
                if account_type == 'PL':
                    # PL 계정은 수익/비용 구분
                    if '매출' in sheet.title:
                        credit = 100000 * (i + 1)
                        debit = 0
                    else:
                        debit = 80000 * (i + 1)
                        credit = 0
                
                balance += debit - credit
                
                sheet.cell(row=row, column=1).value = date_str
                sheet.cell(row=row, column=2).value = f"테스트거래{i+1}"
                sheet.cell(row=row, column=3).value = f"거래처{i+1}"
                sheet.cell(row=row, column=4).value = f"관리{row}"
                sheet.cell(row=row, column=5).value = debit if debit > 0 else ""
                sheet.cell(row=row, column=6).value = credit if credit > 0 else ""
                sheet.cell(row=row, column=7).value = balance
                
                row += 1

class TestLoggingSystem(unittest.TestCase):
    """로깅 시스템 테스트"""
    
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.logger = UTF8LoggingSystem(self.temp_dir)
    
    def test_utf8_encoding_guaranteed(self):
        """UTF-8 인코딩 보장 테스트"""
        # 한글 텍스트 로깅
        korean_text = "한글 테스트 메시지"
        self.logger.log_validation_event(
            'INFO', '10300', 'TEST_EVENT', korean_text
        )
        
        # UTF-8 인코딩 검증
        encoding_check = self.logger.validate_utf8_encoding()
        self.assertTrue(encoding_check['all_files_utf8'])
        
    def test_log_file_creation(self):
        """로그 파일 생성 테스트"""
        log_files = self.logger.get_log_files_info()
        
        for log_type, log_path in log_files['log_files'].items():
            self.assertTrue(os.path.exists(log_path))
            
            # UTF-8로 읽기 가능한지 확인
            with open(log_path, 'r', encoding='utf-8') as f:
                content = f.read()
                self.assertIn('UTF-8', content)
    
    def test_session_summary_creation(self):
        """세션 요약 생성 테스트"""
        summary = self.logger.create_session_summary()
        
        self.assertIn('session_id', summary)
        self.assertIn('statistics', summary)
        self.assertIn('log_files', summary)
    
    def tearDown(self):
        self.logger.close_session()

class TestValidationFramework(unittest.TestCase):
    """검증 프레임워크 테스트"""
    
    def setUp(self):
        self.validator = DataValidator()
        self.workbook = TestDataCreator.create_sample_ledger_workbook()
    
    def test_account_code_extraction(self):
        """계정 코드 추출 테스트"""
        # 정상적인 계정 코드 추출
        account_code = self.validator.extract_account_code("보통예금(10300)")
        self.assertEqual(account_code, "10300")
        
        # 잘못된 형식
        invalid_code = self.validator.extract_account_code("잘못된형식")
        self.assertIsNone(invalid_code)
    
    def test_account_classification(self):
        """계정 분류 테스트"""
        # BS 계정
        bs_type = self.validator.classify_account_type("10300")
        self.assertEqual(bs_type, "BS")
        
        # PL 계정  
        pl_type = self.validator.classify_account_type("41000")
        self.assertEqual(pl_type, "PL")
        
        # VAT 계정
        vat_type = self.validator.classify_account_type("13500")
        self.assertEqual(vat_type, "VAT")
        
        # 알 수 없는 계정
        unknown_type = self.validator.classify_account_type("99999")
        self.assertEqual(unknown_type, "UNKNOWN")
    
    def test_carry_forward_validation(self):
        """전기이월 검증 테스트"""
        sheet = self.workbook["보통예금(10300)"]
        
        carry_forward = self.validator.validate_carry_forward(sheet, "10300")
        self.assertEqual(carry_forward, 1000000)
    
    def test_data_structure_validation(self):
        """데이터 구조 검증 테스트"""
        valid_sheet = "보통예금(10300)"
        is_valid = self.validator.validate_data_structure(self.workbook, valid_sheet)
        self.assertTrue(is_valid)
    
    def test_cross_contamination_detection(self):
        """교차 오염 감지 테스트"""
        # 정상 데이터
        processed_data = {"10300": {"2025": {"01": 100000}}}
        original_data = {"2025": {"10300": {"01": 100000}}}
        
        contamination = self.validator.detect_cross_contamination(processed_data, original_data)
        self.assertEqual(len(contamination), 0)
        
        # 오염된 데이터
        contaminated_processed = {"10300": {"2025": {"01": 100000}}}
        contaminated_original = {"2025": {"10300": {"01": 0}}}  # 원장에 없는 데이터
        
        contamination = self.validator.detect_cross_contamination(
            contaminated_processed, contaminated_original
        )
        self.assertGreater(len(contamination), 0)

class TestMarkingSystem(unittest.TestCase):
    """마킹 시스템 테스트"""
    
    def setUp(self):
        self.marker = YellowMarkingSystem()
        self.workbook = TestDataCreator.create_sample_ledger_workbook()
    
    def test_uncertain_cell_marking(self):
        """불확실한 셀 마킹 테스트"""
        sheet_name = "보통예금(10300)"
        
        result = self.marker.mark_uncertain_cell(
            self.workbook, sheet_name, "G5", "10300", 
            "테스트이슈", "테스트상세내용", 1000000
        )
        
        self.assertTrue(result['marked'])
        self.assertEqual(result['original_value'], 1000000)
        
        # 셀이 실제로 마킹되었는지 확인
        sheet = self.workbook[sheet_name]
        cell = sheet['G5']
        self.assertEqual(cell.fill.start_color.rgb, "00FFFF00")  # 노란색
        self.assertIsNone(cell.value)  # 값이 비워졌는지 확인
    
    def test_marking_summary_creation(self):
        """마킹 요약 생성 테스트"""
        # 몇 개 셀 마킹
        self.marker.mark_uncertain_cell(
            self.workbook, "보통예금(10300)", "G5", "10300", "테스트1", "상세1"
        )
        self.marker.mark_uncertain_cell(
            self.workbook, "매출(41000)", "G10", "41000", "테스트2", "상세2"
        )
        
        # 요약 시트 생성
        result = self.marker.create_marking_summary_sheet(self.workbook)
        
        self.assertTrue(result['summary_created'])
        self.assertEqual(result['total_marked'], 2)
        self.assertIn('마킹요약', self.workbook.sheetnames)
    
    def test_marking_statistics(self):
        """마킹 통계 테스트"""
        # 마킹 수행
        self.marker.mark_uncertain_cell(
            self.workbook, "보통예금(10300)", "G5", "10300", "전기이월불확실", "상세"
        )
        
        stats = self.marker.get_marking_summary()
        
        self.assertEqual(stats['total_marked_cells'], 1)
        self.assertIn('전기이월불확실', stats['marking_statistics']['by_issue_type'])

class TestMCPInterface(unittest.TestCase):
    """MCP 인터페이스 테스트"""
    
    def setUp(self):
        temp_dir = tempfile.mkdtemp()
        self.logger = UTF8LoggingSystem(temp_dir)
        self.mcp = MCPInterface(self.logger)
    
    def test_account_pattern_analysis(self):
        """계정 패턴 분석 테스트"""
        test_data = {"01": 100000, "02": 150000, "03": 200000}
        
        response = self.mcp.analyze_account_patterns(test_data, "BS")
        
        self.assertTrue(response.success)
        self.assertIn('pattern_identified', response.analysis_result)
        self.assertIsNotNone(response.confidence_level)
    
    def test_vat_verification(self):
        """VAT 검증 테스트"""
        transaction_data = [
            {'amount': 110000, 'description': '상품구매', 'counterpart': '공급업체A'},
            {'amount': 50000, 'description': '법무용역', 'counterpart': '법무법인B'}
        ]
        
        response = self.mcp.verify_vat_status(transaction_data)
        
        self.assertTrue(response.success)
        self.assertIn('vat_analysis', response.analysis_result)
    
    def test_batch_size_optimization(self):
        """배치 크기 최적화 테스트"""
        # 소량 데이터
        small_batch_size = self.mcp._determine_optimal_batch_size({"01": 1000}, "BS")
        self.assertGreaterEqual(small_batch_size, 1)
        
        # 대량 데이터 시뮬레이션
        large_data = {f"{i:02d}": 1000 for i in range(1, 25)}  # 24개월 데이터
        large_batch_size = self.mcp._determine_optimal_batch_size(large_data, "PL")
        self.assertEqual(large_batch_size, 1)  # 대량 데이터는 작은 배치
    
    def tearDown(self):
        self.logger.close_session()

class TestBatchProcessor(unittest.TestCase):
    """배치 처리 시스템 테스트"""
    
    def setUp(self):
        temp_dir = tempfile.mkdtemp()
        self.logger = UTF8LoggingSystem(temp_dir)
        self.mcp = MCPInterface(self.logger)
        self.batch_processor = BatchProcessor(self.logger, self.mcp)
    
    def test_batch_task_creation(self):
        """배치 작업 생성 테스트"""
        test_data = {"01": 100000, "02": 200000}
        
        task_id = self.batch_processor.add_batch_task(
            "10300", "BS", "account_pattern_analysis", test_data
        )
        
        self.assertIsNotNone(task_id)
        self.assertIn("10300", task_id)
    
    def test_optimized_batch_creation(self):
        """최적화된 배치 생성 테스트"""
        monthly_data = {f"{i:02d}": 1000 * i for i in range(1, 13)}
        
        task_ids = self.batch_processor.create_optimized_batches(
            "10300", "BS", monthly_data, "account_pattern_analysis"
        )
        
        self.assertGreater(len(task_ids), 0)
    
    def test_batch_statistics(self):
        """배치 통계 테스트"""
        # 작업 추가
        self.batch_processor.add_batch_task(
            "10300", "BS", "account_pattern_analysis", {"01": 1000}
        )
        
        stats = self.batch_processor.batch_stats
        self.assertEqual(stats['total_batches'], 1)
    
    def tearDown(self):
        self.batch_processor.shutdown()
        self.logger.close_session()

class TestMainProcessor(unittest.TestCase):
    """메인 프로세서 통합 테스트"""
    
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file_path = os.path.join(self.temp_dir, "test_ledger.xlsx")
        
        # 테스트용 워크북 생성 및 저장
        workbook = TestDataCreator.create_sample_ledger_workbook()
        workbook.save(self.test_file_path)
    
    def test_full_processing_pipeline(self):
        """전체 처리 파이프라인 테스트"""
        processor = MainProcessor()
        
        result = processor.process_ledger_file(self.test_file_path)
        
        # 기본 결과 검증
        self.assertTrue(result['processing_successful'])
        self.assertIn('session_id', result)
        self.assertIn('output_file', result)
        self.assertGreater(result['processing_time_seconds'], 0)
        
        # 출력 파일 존재 확인
        self.assertTrue(os.path.exists(result['output_file']))
    
    def test_phase_execution(self):
        """각 Phase 실행 테스트"""
        processor = MainProcessor()
        workbook = TestDataCreator.create_sample_ledger_workbook()
        
        # Phase 1 테스트
        phase1_result = processor._execute_phase1_python_basic_validation(workbook)
        self.assertIn('validated_accounts', phase1_result)
        self.assertIn('invalid_sheets', phase1_result)
        
        # 검증된 계정이 있는지 확인
        self.assertGreater(len(phase1_result['validated_accounts']), 0)
    
    def test_error_handling(self):
        """오류 처리 테스트"""
        processor = MainProcessor()
        
        # 존재하지 않는 파일
        with self.assertRaises(FileNotFoundError):
            processor.process_ledger_file("nonexistent_file.xlsx")
    
    def test_config_loading(self):
        """설정 로드 테스트"""
        # 기본 설정
        processor1 = MainProcessor()
        self.assertIn('processing', processor1.config)
        
        # 커스텀 설정 파일
        config_file = os.path.join(self.temp_dir, "test_config.json")
        test_config = {
            'processing': {
                'enable_mcp_analysis': False,
                'timeout_seconds': 600
            }
        }
        
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(test_config, f)
        
        processor2 = MainProcessor(config_file)
        self.assertFalse(processor2.config['processing']['enable_mcp_analysis'])

class TestCLAUDEMdCompliance(unittest.TestCase):
    """CLAUDE.md 규칙 준수 테스트"""
    
    def test_data_integrity_principle(self):
        """데이터 완결성 원칙 테스트"""
        validator = DataValidator()
        marker = YellowMarkingSystem()
        workbook = TestDataCreator.create_sample_ledger_workbook()
        
        # 불확실한 데이터 처리 시뮬레이션
        sheet = workbook["보통예금(10300)"]
        sheet['B5'].value = "전기이월아님"  # 잘못된 전기이월 표시
        
        carry_forward = validator.validate_carry_forward(sheet, "10300")
        
        # 불확실한 경우 None 반환하는지 확인
        self.assertIsNone(carry_forward)
        
        # 노란색 마킹 확인
        self.assertGreater(len(validator.yellow_marks), 0)
    
    def test_utf8_encoding_compliance(self):
        """UTF-8 인코딩 준수 테스트"""
        temp_dir = tempfile.mkdtemp()
        logger = UTF8LoggingSystem(temp_dir)
        
        # 한글 텍스트 처리
        korean_messages = [
            "한글 메시지 테스트",
            "계정코드 추출 실패",
            "전기이월 검증 오류",
            "노란색 마킹 적용"
        ]
        
        for message in korean_messages:
            logger.log_validation_event('INFO', '테스트', '한글테스트', message)
        
        # UTF-8 인코딩 검증
        encoding_check = logger.validate_utf8_encoding()
        self.assertTrue(encoding_check['all_files_utf8'])
        
        logger.close_session()
    
    def test_no_estimation_rule(self):
        """추정값 입력 금지 규칙 테스트"""
        marker = YellowMarkingSystem()
        workbook = TestDataCreator.create_sample_ledger_workbook()
        
        # 불확실한 셀 마킹
        result = marker.mark_uncertain_cell(
            workbook, "보통예금(10300)", "G5", "10300",
            "데이터불확실", "검증불가", 1000000
        )
        
        # 값이 비워졌는지 확인 (추정값 입력 금지)
        sheet = workbook["보통예금(10300)"]
        cell = sheet['G5']
        self.assertIsNone(cell.value)
        
        # 원본값은 기록에 남아있는지 확인
        self.assertEqual(result['original_value'], 1000000)
    
    def test_contamination_prevention(self):
        """교차 오염 방지 테스트"""
        validator = DataValidator()
        
        # 정상적인 경우
        clean_processed = {"10300": {"2025": {"01": 1000}}}
        clean_original = {"2025": {"10300": {"01": 1000}}}
        
        contamination = validator.detect_cross_contamination(clean_processed, clean_original)
        self.assertEqual(len(contamination), 0)
        
        # 교차 오염 케이스
        contaminated = {"10300": {"2025": {"01": 1000}}}
        clean_ledger = {"2025": {"10300": {"01": 0}}}  # 원장에 없는 데이터
        
        contamination = validator.detect_cross_contamination(contaminated, clean_ledger)
        self.assertGreater(len(contamination), 0)
        
        # 오염 감지시 적절한 알림이 있는지 확인
        alert = contamination[0]
        self.assertIn('suspicion', alert)
        self.assertIn('account', alert)

def run_comprehensive_test_suite():
    """종합 테스트 실행"""
    
    print("=== 재무데이터 처리 시스템 종합 테스트 시작 ===")
    print(f"테스트 시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # 테스트 스위트 구성
    test_classes = [
        TestLoggingSystem,
        TestValidationFramework,
        TestMarkingSystem,
        TestMCPInterface,
        TestBatchProcessor,
        TestMainProcessor,
        TestCLAUDEMdCompliance
    ]
    
    total_tests = 0
    passed_tests = 0
    failed_tests = 0
    
    for test_class in test_classes:
        print(f"--- {test_class.__name__} 테스트 시작 ---")
        
        suite = unittest.TestLoader().loadTestsFromTestCase(test_class)
        runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
        result = runner.run(suite)
        
        total_tests += result.testsRun
        passed_tests += result.testsRun - len(result.failures) - len(result.errors)
        failed_tests += len(result.failures) + len(result.errors)
        
        if result.failures:
            print(f"실패한 테스트: {len(result.failures)}개")
            for test, traceback in result.failures:
                print(f"  - {test}: {traceback}")
        
        if result.errors:
            print(f"오류 발생한 테스트: {len(result.errors)}개")
            for test, traceback in result.errors:
                print(f"  - {test}: {traceback}")
        
        print()
    
    print("=== 테스트 결과 요약 ===")
    print(f"총 테스트 수: {total_tests}")
    print(f"성공: {passed_tests}")
    print(f"실패: {failed_tests}")
    print(f"성공률: {(passed_tests/total_tests*100):.1f}%" if total_tests > 0 else "0%")
    print(f"테스트 종료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 테스트 결과를 파일로 저장
    test_report = {
        'timestamp': datetime.now().isoformat(),
        'total_tests': total_tests,
        'passed_tests': passed_tests,
        'failed_tests': failed_tests,
        'success_rate': (passed_tests/total_tests*100) if total_tests > 0 else 0,
        'test_classes': [cls.__name__ for cls in test_classes]
    }
    
    os.makedirs('./log', exist_ok=True)
    report_file = f"./log/test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    with open(report_file, 'w', encoding='utf-8') as f:
        json.dump(test_report, f, ensure_ascii=False, indent=2)
    
    print(f"테스트 보고서 저장: {report_file}")
    
    return failed_tests == 0

if __name__ == "__main__":
    success = run_comprehensive_test_suite()
    sys.exit(0 if success else 1)
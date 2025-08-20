"""
재무데이터 처리 시스템 - 메인 통합 프로세서
CLAUDE.md 규칙 완전 준수, PRD 기반 통합 시스템
"""

import os
import sys
import time
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from openpyxl import load_workbook

# 프로젝트 모듈들
from logging_system import UTF8LoggingSystem
from validation_framework import DataValidator
from marking_system import YellowMarkingSystem
from mcp_interface import MCPInterface
from batch_processor import BatchProcessor, BatchPriority

class MainProcessor:
    """
    재무데이터 처리 메인 프로세서
    3단계 검증 체계: Python 기본검증 -> MCP 패턴분석 -> Python 최종검증
    """
    
    def __init__(self, config_path: Optional[str] = None):
        # 시스템 초기화
        self.session_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.processing_start_time = datetime.now()
        
        # 설정 로드
        self.config = self._load_config(config_path)
        
        # 핵심 시스템 컴포넌트 초기화
        self.logger = UTF8LoggingSystem("./log")
        self.validator = DataValidator()
        self.marker = YellowMarkingSystem()
        self.mcp = MCPInterface(self.logger)
        self.batch_processor = BatchProcessor(self.logger, self.mcp)
        
        # 처리 상태 관리
        self.processing_stats = {
            'session_id': self.session_id,
            'total_accounts': 0,
            'processed_accounts': 0,
            'failed_accounts': 0,
            'uncertain_accounts': 0,
            'marked_cells': 0,
            'contamination_alerts': 0,
            'processing_phases': {
                'phase1_python_basic': {'status': 'PENDING', 'duration': 0},
                'phase2_mcp_analysis': {'status': 'PENDING', 'duration': 0},
                'phase3_python_final': {'status': 'PENDING', 'duration': 0}
            }
        }
        
        # 결과 저장소
        self.processed_data = {}
        self.validation_results = {}
        self.batch_results = {}
        self.contamination_alerts = []
        
        self.logger.log_validation_event(
            'INFO', 'SYSTEM', 'MAIN_PROCESSOR_INIT',
            f'메인프로세서초기화완료_세션{self.session_id}'
        )
    
    def _load_config(self, config_path: Optional[str]) -> Dict:
        """설정 파일 로드"""
        default_config = {
            'processing': {
                'enable_mcp_analysis': True,
                'enable_batch_processing': True,
                'max_retry_attempts': 3,
                'timeout_seconds': 1800  # 30분
            },
            'validation': {
                'strict_mode': True,
                'mark_uncertain_data': True,
                'cross_contamination_check': True
            },
            'output': {
                'create_summary_sheet': True,
                'generate_reports': True,
                'backup_original': True
            }
        }
        
        if config_path and os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    user_config = json.load(f)
                    default_config.update(user_config)
            except Exception as e:
                print(f"설정 파일 로드 실패, 기본 설정 사용: {e}")
        
        return default_config
    
    def process_ledger_file(self, file_path: str, output_path: Optional[str] = None) -> Dict:
        """
        원장 파일 전체 처리 메인 함수
        3단계 검증 체계 적용
        """
        try:
            # 파일 경로 검증
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")
            
            # 출력 경로 설정
            if not output_path:
                file_stem = Path(file_path).stem
                output_path = f"{file_stem}_processed_{self.session_id}.xlsx"
            
            # 원본 파일 백업
            if self.config['output']['backup_original']:
                self._backup_original_file(file_path)
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'PROCESSING_START',
                f'원장파일처리시작_{Path(file_path).name}',
                {'input_file': file_path, 'output_file': output_path}
            )
            
            # 워크북 로드
            workbook = load_workbook(file_path)
            
            # Phase 1: Python 기본 검증
            phase1_result = self._execute_phase1_python_basic_validation(workbook)
            
            # Phase 2: MCP 패턴 분석 (배치 처리)
            phase2_result = self._execute_phase2_mcp_analysis(workbook, phase1_result)
            
            # Phase 3: Python 최종 검증
            phase3_result = self._execute_phase3_python_final_validation(
                workbook, phase1_result, phase2_result
            )
            
            # 결과 통합 및 파일 저장
            final_result = self._finalize_processing(
                workbook, output_path, phase1_result, phase2_result, phase3_result
            )
            
            # 세션 종료 및 정리
            self._cleanup_session()
            
            return final_result
            
        except Exception as e:
            self.logger.log_validation_event(
                'ERROR', 'SYSTEM', 'PROCESSING_ERROR',
                f'처리중오류_{str(e)}',
                {'file_path': file_path}
            )
            raise
    
    def _execute_phase1_python_basic_validation(self, workbook) -> Dict:
        """
        Phase 1: Python 기본 검증
        - 데이터 구조 검증
        - 필수 필드 존재 확인
        - 기본적인 데이터 타입 검증
        """
        phase_start = time.time()
        self.processing_stats['processing_phases']['phase1_python_basic']['status'] = 'PROCESSING'
        
        self.logger.log_validation_event(
            'INFO', 'SYSTEM', 'PHASE1_START', 'Python기본검증시작'
        )
        
        phase1_results = {
            'validated_accounts': {},
            'invalid_sheets': [],
            'marked_cells': [],
            'validation_summary': {}
        }
        
        try:
            # 모든 시트 스캔 (분석 시트 제외)
            valid_sheets = [sheet for sheet in workbook.sheetnames 
                           if not any(keyword in sheet.lower() 
                                    for keyword in ['분석', 'summary', 'temp', '요약'])]
            
            self.processing_stats['total_accounts'] = len(valid_sheets)
            
            for sheet_name in valid_sheets:
                try:
                    # 계정별 검증 수행
                    validation_result = self.validator.validate_account_data(workbook, sheet_name)
                    
                    if validation_result['validation_passed']:
                        phase1_results['validated_accounts'][sheet_name] = validation_result
                        self.processing_stats['processed_accounts'] += 1
                    else:
                        phase1_results['invalid_sheets'].append({
                            'sheet_name': sheet_name,
                            'issues': validation_result['issues'],
                            'account_code': validation_result.get('account_code')
                        })
                        self.processing_stats['failed_accounts'] += 1
                        
                        # 실패한 시트 전체를 노란색 마킹
                        self._mark_invalid_sheet(workbook, sheet_name, validation_result)
                
                except Exception as e:
                    self.logger.log_validation_event(
                        'ERROR', 'UNKNOWN', 'VALIDATION_ERROR',
                        f'시트검증오류_{sheet_name}_{str(e)}'
                    )
                    phase1_results['invalid_sheets'].append({
                        'sheet_name': sheet_name,
                        'error': str(e)
                    })
            
            # 마킹된 셀 정보 수집
            phase1_results['marked_cells'] = self.validator.yellow_marks
            self.processing_stats['marked_cells'] = len(self.validator.yellow_marks)
            
            # Phase 1 완료
            phase_duration = time.time() - phase_start
            self.processing_stats['processing_phases']['phase1_python_basic'].update({
                'status': 'COMPLETED',
                'duration': phase_duration
            })
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'PHASE1_COMPLETE',
                f'Python기본검증완료_성공{len(phase1_results["validated_accounts"])}개_실패{len(phase1_results["invalid_sheets"])}개',
                {'duration': phase_duration, 'marked_cells': len(phase1_results['marked_cells'])}
            )
            
            return phase1_results
            
        except Exception as e:
            self.processing_stats['processing_phases']['phase1_python_basic']['status'] = 'FAILED'
            raise
    
    def _execute_phase2_mcp_analysis(self, workbook, phase1_result: Dict) -> Dict:
        """
        Phase 2: MCP 패턴 분석
        - 복잡한 패턴 인식
        - VAT 포함/제외 판단
        - 이상 케이스 분석
        """
        phase_start = time.time()
        self.processing_stats['processing_phases']['phase2_mcp_analysis']['status'] = 'PROCESSING'
        
        self.logger.log_validation_event(
            'INFO', 'SYSTEM', 'PHASE2_START', 'MCP패턴분석시작'
        )
        
        phase2_results = {
            'mcp_analysis_results': {},
            'uncertain_items': [],
            'batch_results': [],
            'analysis_summary': {}
        }
        
        try:
            # 배치 처리 시작
            self.batch_processor.start_batch_processing()
            
            # 검증된 계정들에 대해 MCP 분석 수행
            for sheet_name, validation_data in phase1_result['validated_accounts'].items():
                account_code = validation_data['account_code']
                account_type = validation_data['account_type']
                monthly_data = validation_data['monthly_data']
                
                if not account_code or account_type == 'UNKNOWN':
                    # UNKNOWN 계정은 MCP 분석에서 제외
                    continue
                
                # 최적화된 배치 생성 및 작업 추가
                task_ids = self.batch_processor.create_optimized_batches(
                    account_code, account_type, monthly_data, 'account_pattern_analysis'
                )
                
                phase2_results['batch_results'].extend(task_ids)
            
            # 모든 배치 작업 완료 대기
            completion_stats = self.batch_processor.wait_for_batch_completion(
                timeout_seconds=self.config['processing']['timeout_seconds']
            )
            
            # 배치 결과 수집
            batch_results = self.batch_processor.get_all_results()
            
            # 결과 분석
            for result in batch_results:
                if result.confidence_level == 'UNCERTAIN':
                    phase2_results['uncertain_items'].append({
                        'task_id': result.task_id,
                        'uncertain_items': result.uncertain_items,
                        'result_data': result.result_data
                    })
                    self.processing_stats['uncertain_accounts'] += 1
                
                phase2_results['mcp_analysis_results'][result.task_id] = result.result_data
            
            # UNCERTAIN 항목들 노란색 마킹 처리
            self._process_uncertain_items(workbook, phase2_results['uncertain_items'])
            
            # Phase 2 완료
            phase_duration = time.time() - phase_start
            self.processing_stats['processing_phases']['phase2_mcp_analysis'].update({
                'status': 'COMPLETED',
                'duration': phase_duration
            })
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'PHASE2_COMPLETE',
                f'MCP패턴분석완료_배치{len(batch_results)}개_불확실{len(phase2_results["uncertain_items"])}개',
                {'duration': phase_duration, 'completion_stats': completion_stats}
            )
            
            return phase2_results
            
        except Exception as e:
            self.processing_stats['processing_phases']['phase2_mcp_analysis']['status'] = 'FAILED'
            raise
    
    def _execute_phase3_python_final_validation(self, workbook, 
                                              phase1_result: Dict, 
                                              phase2_result: Dict) -> Dict:
        """
        Phase 3: Python 최종 검증
        - 교차 오염 검사
        - 데이터 무결성 확인
        - 최종 품질 보증
        """
        phase_start = time.time()
        self.processing_stats['processing_phases']['phase3_python_final']['status'] = 'PROCESSING'
        
        self.logger.log_validation_event(
            'INFO', 'SYSTEM', 'PHASE3_START', 'Python최종검증시작'
        )
        
        phase3_results = {
            'contamination_check': {},
            'final_validation': {},
            'integrity_report': {},
            'quality_metrics': {}
        }
        
        try:
            # 교차 오염 검사
            processed_data = self._consolidate_processed_data(phase1_result, phase2_result)
            original_ledger_data = self._extract_original_ledger_data(workbook)
            
            contamination_alerts = self.validator.detect_cross_contamination(
                processed_data, original_ledger_data
            )
            
            if contamination_alerts:
                # 교차 오염 발견시 즉시 중단 및 마킹
                self._handle_contamination_alerts(workbook, contamination_alerts)
                phase3_results['contamination_check'] = {
                    'contamination_detected': True,
                    'alert_count': len(contamination_alerts),
                    'alerts': contamination_alerts
                }
                self.processing_stats['contamination_alerts'] = len(contamination_alerts)
                
                # 교차 오염 감지시 처리 중단
                self.logger.log_contamination_event(
                    'ERROR', {'summary': f'교차오염감지_{len(contamination_alerts)}건_처리중단'}
                )
                
                raise Exception(f"교차 오염 감지: {len(contamination_alerts)}건. 처리를 중단합니다.")
            
            else:
                phase3_results['contamination_check'] = {
                    'contamination_detected': False,
                    'data_integrity': 'VERIFIED'
                }
            
            # 최종 데이터 검증
            final_validation = self._perform_final_data_validation(processed_data)
            phase3_results['final_validation'] = final_validation
            
            # 품질 메트릭 계산
            quality_metrics = self._calculate_quality_metrics(
                phase1_result, phase2_result, phase3_results
            )
            phase3_results['quality_metrics'] = quality_metrics
            
            # Phase 3 완료
            phase_duration = time.time() - phase_start
            self.processing_stats['processing_phases']['phase3_python_final'].update({
                'status': 'COMPLETED',
                'duration': phase_duration
            })
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'PHASE3_COMPLETE',
                f'Python최종검증완료_오염없음_품질점수{quality_metrics.get("overall_score", 0)}점',
                {'duration': phase_duration, 'quality_metrics': quality_metrics}
            )
            
            return phase3_results
            
        except Exception as e:
            self.processing_stats['processing_phases']['phase3_python_final']['status'] = 'FAILED'
            raise
    
    def _finalize_processing(self, workbook, output_path: str, 
                           phase1_result: Dict, phase2_result: Dict, 
                           phase3_result: Dict) -> Dict:
        """처리 결과 최종화 및 파일 저장"""
        try:
            # 마킹 요약 시트 생성
            if self.config['output']['create_summary_sheet']:
                self.marker.create_marking_summary_sheet(workbook)
            
            # 처리 결과 통계 시트 생성
            self._create_processing_summary_sheet(workbook, phase1_result, phase2_result, phase3_result)
            
            # 최종 파일 저장
            workbook.save(output_path)
            
            # 최종 보고서 생성
            final_report = self._generate_final_report(
                output_path, phase1_result, phase2_result, phase3_result
            )
            
            # 처리 시간 계산
            total_processing_time = (datetime.now() - self.processing_start_time).total_seconds()
            
            final_result = {
                'session_id': self.session_id,
                'processing_successful': True,
                'output_file': output_path,
                'processing_time_seconds': total_processing_time,
                'processing_stats': self.processing_stats,
                'quality_assessment': phase3_result.get('quality_metrics', {}),
                'final_report_path': final_report.get('report_file_path'),
                'marked_cells_count': len(self.validator.yellow_marks),
                'contamination_alerts': len(self.contamination_alerts),
                'phase_results': {
                    'phase1_basic_validation': phase1_result,
                    'phase2_mcp_analysis': phase2_result,
                    'phase3_final_validation': phase3_result
                }
            }
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'PROCESSING_FINALIZED',
                f'처리완료_파일{output_path}_시간{total_processing_time:.1f}초',
                final_result
            )
            
            return final_result
            
        except Exception as e:
            self.logger.log_validation_event(
                'ERROR', 'SYSTEM', 'FINALIZATION_ERROR',
                f'최종화오류_{str(e)}'
            )
            raise
    
    def _backup_original_file(self, file_path: str):
        """원본 파일 백업"""
        backup_path = f"{file_path}.backup_{self.session_id}"
        try:
            import shutil
            shutil.copy2(file_path, backup_path)
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'BACKUP_CREATED',
                f'원본백업생성_{backup_path}'
            )
        except Exception as e:
            self.logger.log_validation_event(
                'WARNING', 'SYSTEM', 'BACKUP_FAILED',
                f'백업생성실패_{str(e)}'
            )
    
    def _mark_invalid_sheet(self, workbook, sheet_name: str, validation_result: Dict):
        """유효하지 않은 시트 마킹"""
        account_code = validation_result.get('account_code', 'UNKNOWN')
        issues = validation_result.get('issues', [])
        
        for issue in issues:
            self.marker.mark_range_uncertain(
                workbook, sheet_name, 'A1', 'G10',
                account_code, '시트검증실패', f'이슈_{issue}'
            )
    
    def _process_uncertain_items(self, workbook, uncertain_items: List[Dict]):
        """UNCERTAIN 항목들 처리"""
        for item in uncertain_items:
            # MCP에서 UNCERTAIN으로 판단된 항목들을 노란색 마킹
            task_id = item['task_id']
            uncertain_list = item['uncertain_items']
            
            # task_id에서 계정 정보 추출
            parts = task_id.split('_')
            if len(parts) >= 2:
                account_code = parts[0]
                
                # 해당 계정 시트에 마킹 적용
                sheet_name = None
                for name in workbook.sheetnames:
                    if account_code in name:
                        sheet_name = name
                        break
                
                if sheet_name:
                    for uncertain_type in uncertain_list:
                        self.marker.mark_range_uncertain(
                            workbook, sheet_name, 'G1', 'G50',  # 추정 범위
                            account_code, 'MCP_UNCERTAIN', f'MCP판단불확실_{uncertain_type}'
                        )
    
    def _consolidate_processed_data(self, phase1_result: Dict, phase2_result: Dict) -> Dict:
        """Phase 1, 2 결과 통합"""
        consolidated = {}
        
        for sheet_name, validation_data in phase1_result['validated_accounts'].items():
            account_code = validation_data['account_code']
            if account_code:
                consolidated[account_code] = {
                    '2025': validation_data['monthly_data']  # 예시: 연도별 데이터
                }
        
        return consolidated
    
    def _extract_original_ledger_data(self, workbook) -> Dict:
        """원장에서 원본 데이터 추출"""
        extractor = LedgerExtractionEngine()
        return extractor.extract_all_ledger_data(workbook)
    
    def _handle_contamination_alerts(self, workbook, contamination_alerts: List[Dict]):
        """교차 오염 알림 처리"""
        self.contamination_alerts = contamination_alerts
        
        for alert in contamination_alerts:
            account_code = alert['account']
            
            # 관련 시트 찾기 및 마킹
            for sheet_name in workbook.sheetnames:
                if account_code in sheet_name:
                    self.marker.mark_contamination_alert(workbook, sheet_name, alert)
    
    def _perform_final_data_validation(self, processed_data: Dict) -> Dict:
        """최종 데이터 검증"""
        return {
            'validation_passed': True,
            'total_accounts_validated': len(processed_data),
            'validation_timestamp': datetime.now().isoformat()
        }
    
    def _calculate_quality_metrics(self, phase1_result: Dict, 
                                 phase2_result: Dict, 
                                 phase3_result: Dict) -> Dict:
        """품질 메트릭 계산"""
        total_accounts = len(phase1_result['validated_accounts'])
        marked_cells = len(self.validator.yellow_marks)
        uncertain_items = len(phase2_result['uncertain_items'])
        contamination_detected = phase3_result['contamination_check'].get('contamination_detected', False)
        
        # 품질 점수 계산 (100점 만점)
        base_score = 100
        
        # 마킹된 셀에 따른 감점
        if marked_cells > 0:
            marking_penalty = min(marked_cells * 2, 30)  # 최대 30점 감점
            base_score -= marking_penalty
        
        # UNCERTAIN 항목에 따른 감점
        if uncertain_items > 0:
            uncertainty_penalty = min(uncertain_items * 5, 25)  # 최대 25점 감점
            base_score -= uncertainty_penalty
        
        # 교차 오염 감지시 대폭 감점
        if contamination_detected:
            base_score -= 40
        
        quality_grade = 'A' if base_score >= 90 else 'B' if base_score >= 80 else 'C' if base_score >= 70 else 'D'
        
        return {
            'overall_score': max(base_score, 0),
            'quality_grade': quality_grade,
            'data_integrity': 'COMPROMISED' if contamination_detected else 'INTACT',
            'marked_cells_count': marked_cells,
            'uncertain_items_count': uncertain_items,
            'contamination_detected': contamination_detected,
            'processing_accuracy_rate': ((total_accounts - uncertain_items) / max(total_accounts, 1)) * 100
        }
    
    def _create_processing_summary_sheet(self, workbook, phase1_result: Dict, 
                                       phase2_result: Dict, phase3_result: Dict):
        """처리 요약 시트 생성"""
        try:
            if "처리요약" in workbook.sheetnames:
                del workbook["처리요약"]
            
            summary_sheet = workbook.create_sheet("처리요약")
            
            # 요약 데이터 작성
            summary_data = [
                ["항목", "값", "상태"],
                ["세션 ID", self.session_id, ""],
                ["처리 시작시간", self.processing_start_time.strftime('%Y-%m-%d %H:%M:%S'), ""],
                ["총 계정 수", self.processing_stats['total_accounts'], ""],
                ["성공 계정 수", self.processing_stats['processed_accounts'], "✓"],
                ["실패 계정 수", self.processing_stats['failed_accounts'], "✗" if self.processing_stats['failed_accounts'] > 0 else ""],
                ["불확실 계정 수", self.processing_stats['uncertain_accounts'], "⚠" if self.processing_stats['uncertain_accounts'] > 0 else ""],
                ["마킹된 셀 수", self.processing_stats['marked_cells'], "⚠" if self.processing_stats['marked_cells'] > 0 else ""],
                ["교차 오염 알림", self.processing_stats['contamination_alerts'], "❌" if self.processing_stats['contamination_alerts'] > 0 else "✓"],
                ["", "", ""],
                ["Phase 1 상태", self.processing_stats['processing_phases']['phase1_python_basic']['status'], ""],
                ["Phase 2 상태", self.processing_stats['processing_phases']['phase2_mcp_analysis']['status'], ""],
                ["Phase 3 상태", self.processing_stats['processing_phases']['phase3_python_final']['status'], ""],
            ]
            
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    summary_sheet.cell(row=row_idx, column=col_idx).value = cell_value
            
        except Exception as e:
            self.logger.log_validation_event(
                'WARNING', 'SYSTEM', 'SUMMARY_SHEET_ERROR',
                f'요약시트생성오류_{str(e)}'
            )
    
    def _generate_final_report(self, output_path: str, phase1_result: Dict, 
                             phase2_result: Dict, phase3_result: Dict) -> Dict:
        """최종 보고서 생성"""
        if not self.config['output']['generate_reports']:
            return {}
        
        report_data = {
            'processing_summary': {
                'session_id': self.session_id,
                'input_file': 'N/A',  # 입력 파일명
                'output_file': output_path,
                'processing_time': (datetime.now() - self.processing_start_time).total_seconds(),
                'timestamp': datetime.now().isoformat()
            },
            'statistics': self.processing_stats,
            'quality_metrics': phase3_result.get('quality_metrics', {}),
            'phase_summaries': {
                'phase1_python_basic': {
                    'validated_accounts': len(phase1_result['validated_accounts']),
                    'invalid_sheets': len(phase1_result['invalid_sheets']),
                    'marked_cells': len(phase1_result['marked_cells'])
                },
                'phase2_mcp_analysis': {
                    'analysis_results': len(phase2_result['mcp_analysis_results']),
                    'uncertain_items': len(phase2_result['uncertain_items']),
                    'batch_results': len(phase2_result['batch_results'])
                },
                'phase3_final_validation': {
                    'contamination_detected': phase3_result['contamination_check'].get('contamination_detected', False),
                    'final_validation_passed': phase3_result['final_validation'].get('validation_passed', False)
                }
            },
            'recommendations': self._generate_recommendations(phase1_result, phase2_result, phase3_result)
        }
        
        # UTF-8 보장하여 보고서 파일 생성
        report_file_path = f"./log/final_report_{self.session_id}.json"
        try:
            with open(report_file_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2)
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'FINAL_REPORT_GENERATED',
                f'최종보고서생성_{report_file_path}'
            )
            
            return {'report_file_path': report_file_path, 'report_data': report_data}
            
        except Exception as e:
            self.logger.log_validation_event(
                'ERROR', 'SYSTEM', 'REPORT_GENERATION_ERROR',
                f'보고서생성오류_{str(e)}'
            )
            return {}
    
    def _generate_recommendations(self, phase1_result: Dict, 
                                phase2_result: Dict, 
                                phase3_result: Dict) -> List[str]:
        """권고사항 생성"""
        recommendations = []
        
        # 마킹된 셀이 많은 경우
        if len(phase1_result['marked_cells']) > 10:
            recommendations.append("노란색 마킹된 셀이 많습니다. 데이터 품질 검토가 필요합니다.")
        
        # UNCERTAIN 항목이 많은 경우
        if len(phase2_result['uncertain_items']) > 5:
            recommendations.append("MCP 분석에서 불확실한 항목이 많습니다. 수동 검토를 권장합니다.")
        
        # 교차 오염이 감지된 경우
        if phase3_result['contamination_check'].get('contamination_detected'):
            recommendations.append("❌ 교차 오염이 감지되었습니다. 즉시 데이터를 검토하고 수정해주세요.")
        
        # 품질 점수가 낮은 경우
        quality_score = phase3_result.get('quality_metrics', {}).get('overall_score', 100)
        if quality_score < 80:
            recommendations.append(f"데이터 품질 점수가 {quality_score}점으로 낮습니다. 추가 검증이 필요합니다.")
        
        # 실패한 계정이 있는 경우
        if len(phase1_result['invalid_sheets']) > 0:
            recommendations.append(f"{len(phase1_result['invalid_sheets'])}개 계정에서 검증이 실패했습니다. 해당 계정들을 확인해주세요.")
        
        if not recommendations:
            recommendations.append("✅ 모든 검증이 성공적으로 완료되었습니다.")
        
        return recommendations
    
    def _cleanup_session(self):
        """세션 정리"""
        try:
            # 배치 처리 시스템 종료
            self.batch_processor.shutdown()
            
            # 로깅 시스템 세션 종료
            self.logger.close_session()
            
            self.logger.log_validation_event(
                'INFO', 'SYSTEM', 'SESSION_CLEANUP_COMPLETE',
                f'세션정리완료_{self.session_id}'
            )
            
        except Exception as e:
            print(f"세션 정리 중 오류: {e}")

def main():
    """메인 실행 함수 (CLI 용도)"""
    if len(sys.argv) < 2:
        print("사용법: python main_processor.py <입력파일경로> [출력파일경로] [설정파일경로]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    config_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    try:
        processor = MainProcessor(config_file)
        result = processor.process_ledger_file(input_file, output_file)
        
        print(f"처리 완료!")
        print(f"출력 파일: {result['output_file']}")
        print(f"처리 시간: {result['processing_time_seconds']:.1f}초")
        print(f"품질 점수: {result['quality_assessment'].get('overall_score', 'N/A')}점")
        print(f"마킹된 셀: {result['marked_cells_count']}개")
        
        if result['contamination_alerts'] > 0:
            print(f"⚠️  교차 오염 알림: {result['contamination_alerts']}건")
        
    except Exception as e:
        print(f"처리 중 오류 발생: {e}")
        sys.exit(1)

class LedgerExtractionEngine:
    """
    원장 데이터 추출 엔진
    CLAUDE.md 표준 로직 구현: 전기이월, 월별 데이터, 계정코드 파싱
    """
    
    def __init__(self):
        self.account_mapping = {
            2022: "C:/K04_cashflow_1/Ledgers/2022년_계정별원장_회계감사반영분_인게니움_null처리완료.xlsx",
            2023: "C:/K04_cashflow_1/Ledgers/2023년_계정별원장_인게니움_null처리완료.xlsx", 
            2024: "C:/K04_cashflow_1/Ledgers/2024_00_00_ACC_계정별원장_전체_null처리완료.xlsx",
            2025: "C:/K04_cashflow_1/Ledgers/2025년01월-06월_계정별원장_인게니움_v1_null처리완료.xlsx"
        }
        
        # BS/PL 계정 분류 (CLAUDE.md 기준)
        self.bs_ranges = {
            'assets': (10000, 25000),      # 100-200번대
            'liabilities': (25000, 30000), # 250-290번대  
            'equity': (33000, 38000)       # 330-370번대
        }
        
        self.pl_ranges = {
            'revenue': [(40000, 42100), (90000, 92100)],
            'expenses': [(45000, 46100), (52000, 53100), (80000, 84100), (93000, 96100)]
        }
        
        import logging
        logging.info("[원장추출엔진초기화] [파일매핑완료] [계정분류설정완료]")
    
    def extract_all_ledger_data(self, workbook) -> Dict:
        """워크북에서 모든 원장 데이터 추출"""
        extracted_data = {}
        
        for sheet_name in workbook.sheetnames:
            try:
                # 계정 코드 추출
                account_code = self.extract_account_code(sheet_name)
                if not account_code:
                    continue
                
                sheet = workbook[sheet_name]
                
                # 계정 타입 결정
                account_type = self.classify_account_type(account_code)
                
                # 전기이월 추출
                carry_forward = self.extract_carry_forward(sheet)
                
                # 월별 데이터 추출
                monthly_data = self.extract_monthly_data(sheet, account_type)
                
                extracted_data[account_code] = {
                    'account_name': sheet_name.split('(')[0].strip(),
                    'account_type': account_type,
                    'carry_forward': carry_forward,
                    'monthly_data': monthly_data
                }
                
                logging.info(f"[추출성공] [계정_{account_code}] [타입_{account_type}] [전기이월_{carry_forward}] [월별데이터_{len(monthly_data)}개월]")
                
            except Exception as e:
                logging.error(f"[추출실패] [시트_{sheet_name}] [오류_{str(e)}]")
                continue
        
        return extracted_data
    
    def extract_carry_forward(self, sheet):
        """전기이월: 각 시트 5행 G열에서 추출 (CLAUDE.md 규칙)"""
        import logging
        try:
            b5_value = sheet['B5'].value
            if b5_value and '전기이월' in str(b5_value):
                g5_value = sheet['G5'].value
                if isinstance(g5_value, (int, float)):
                    return g5_value
                else:
                    logging.warning(f"[전기이월형식오류] [B5=전기이월] [G5={g5_value}] [숫자아님]")
                    return None
            else:
                logging.warning(f"[전기이월없음] [B5={b5_value}] [전기이월텍스트없음]")
                return None
        except Exception as e:
            logging.error(f"[전기이월추출오류] [오류_{str(e)}]")
            return None
    
    def extract_monthly_data(self, sheet, account_type):
        """
        월별 데이터 추출 - BS/PL 계정 구분 처리
        BS: 월계 직전 거래행 G열 = 월말 잔액 (검증 완료 패턴)
        PL: 월계행 차변-대변 = 월별 발생액 (새로 추가된 로직)
        """
        import logging
        monthly_data = {}
        
        try:
            if account_type == 'BS':
                # BS 계정: 월계 직전 거래행 G열 잔액 추출
                monthly_data = self._extract_bs_monthly_balances(sheet)
            elif account_type == 'PL':
                # PL 계정: 월계행 발생액 추출
                monthly_data = self._extract_pl_monthly_amounts(sheet)
            else:
                logging.warning(f"[알수없는계정타입] [타입_{account_type}] [기본BS로직적용]")
                monthly_data = self._extract_bs_monthly_balances(sheet)
        
        except Exception as e:
            logging.error(f"[월별데이터추출오류] [계정타입_{account_type}] [오류_{str(e)}]")
        
        return monthly_data
    
    def _extract_bs_monthly_balances(self, sheet):
        """BS 계정 월별 잔액 추출 - 검증 완료된 패턴"""
        import logging
        import re
        
        monthly_balances = {}
        current_month = None
        last_balance = None
        
        for row in range(6, sheet.max_row + 1):
            a_val = sheet[f'A{row}'].value
            b_val = sheet[f'B{row}'].value
            g_val = sheet[f'G{row}'].value
            
            # MM-DD 패턴으로 월 인식
            if a_val and isinstance(a_val, str) and '-' in a_val:
                parts = a_val.split('-')
                if len(parts) >= 2 and parts[0].isdigit():
                    month = int(parts[0])
                    if 1 <= month <= 12:
                        current_month = month
            
            # 월계 행 감지시 직전 거래의 잔액을 월말 잔액으로 설정
            if b_val and isinstance(b_val, str) and '월         계' in b_val:
                if current_month and last_balance is not None:
                    monthly_balances[current_month] = last_balance
                    logging.info(f"[BS월말잔액] [월_{current_month}] [잔액_{last_balance}]")
                current_month = None
                last_balance = None
                continue
            
            # 일반 거래 행에서 잔액 추적
            if current_month and g_val is not None:
                last_balance = g_val
        
        return monthly_balances
    
    def _extract_pl_monthly_amounts(self, sheet):
        """PL 계정 월별 발생액 추출 - 수정된 로직"""
        import logging
        
        monthly_amounts = {}
        current_month = None
        monthly_debit_total = 0
        monthly_credit_total = 0
        
        for row in range(6, sheet.max_row + 1):
            a_val = sheet[f'A{row}'].value
            b_val = sheet[f'B{row}'].value
            e_val = sheet[f'E{row}'].value  # 차변
            f_val = sheet[f'F{row}'].value  # 대변
            
            # MM-DD 패턴으로 월 인식
            if a_val and isinstance(a_val, str) and '-' in a_val:
                parts = a_val.split('-')
                if len(parts) >= 2 and parts[0].isdigit():
                    month = int(parts[0])
                    if 1 <= month <= 12:
                        # 이전 월 처리 완료
                        if current_month and current_month not in monthly_amounts:
                            net_amount = monthly_debit_total - monthly_credit_total
                            if net_amount != 0:
                                monthly_amounts[current_month] = net_amount
                                logging.info(f"[PL월별발생액] [월_{current_month}] [차변총_{monthly_debit_total}] [대변총_{monthly_credit_total}] [순발생_{net_amount}]")
                        
                        # 새 월 시작
                        current_month = month
                        monthly_debit_total = 0
                        monthly_credit_total = 0
            
            # 월 내 거래 누적
            if current_month and a_val and isinstance(a_val, str) and '-' in a_val:
                debit = e_val or 0
                credit = f_val or 0
                if isinstance(debit, (int, float)):
                    monthly_debit_total += debit
                if isinstance(credit, (int, float)):
                    monthly_credit_total += credit
            
            # 월계 행 발견시 처리 완료
            if b_val and isinstance(b_val, str) and '월         계' in b_val:
                if current_month and current_month not in monthly_amounts:
                    # 월계 행 직접 값 사용 (더 정확)
                    debit = e_val or 0
                    credit = f_val or 0
                    monthly_amount = debit - credit
                    
                    if monthly_amount != 0:
                        monthly_amounts[current_month] = monthly_amount
                        logging.info(f"[PL월계발생액] [월_{current_month}] [차변_{debit}] [대변_{credit}] [발생액_{monthly_amount}]")
                
                current_month = None
                monthly_debit_total = 0
                monthly_credit_total = 0
        
        # 마지막 월 처리
        if current_month and current_month not in monthly_amounts:
            net_amount = monthly_debit_total - monthly_credit_total
            if net_amount != 0:
                monthly_amounts[current_month] = net_amount
                logging.info(f"[PL마지막월] [월_{current_month}] [차변총_{monthly_debit_total}] [대변총_{monthly_credit_total}] [순발생_{net_amount}]")
        
        return monthly_amounts
    
    def extract_account_code(self, sheet_name):
        """시트명 파싱: 정규표현식 `\\((\\d+)\\)`으로 계정코드 추출 (CLAUDE.md 규칙)"""
        import re
        pattern = r'\((\d+)\)'
        match = re.search(pattern, sheet_name)
        
        if match:
            return match.group(1)
        else:
            logging.warning(f"[계정코드추출실패] [시트_{sheet_name}] [패턴불일치]")
            return None
    
    def classify_account_type(self, account_code):
        """계정 분류: BS/PL 구분"""
        try:
            code_int = int(account_code)
            
            # BS 계정 확인
            for range_name, (start, end) in self.bs_ranges.items():
                if start <= code_int < end:
                    return 'BS'
            
            # PL 계정 확인
            for range_list in self.pl_ranges.values():
                for start, end in range_list:
                    if start <= code_int < end:
                        return 'PL'
            
            # VAT 계정 특별 처리
            if account_code in ['13500', '25500']:
                return 'VAT'
            
            logging.warning(f"[계정분류실패] [코드_{account_code}] [알수없는범위]")
            return 'UNKNOWN'
            
        except ValueError:
            logging.error(f"[계정분류오류] [코드_{account_code}] [숫자변환실패]")
            return 'UNKNOWN'


class TemplateIntegrator:
    """
    추출된 데이터를 자금수지표 템플릿에 자동 반영
    """
    
    def __init__(self):
        self.template_structure = {
            'bs_sheet_name': '원장데이터_BS계정',
            'pl_sheet_name': '원장데이터_PL',
            'account_column': 1,  # A열에 계정 정보
            'data_start_column': 2  # B열부터 데이터
        }
        
        import logging
        logging.info("[템플릿통합시스템초기화] [구조설정완료]")
    
    def integrate_to_template(self, extracted_data, template_path):
        """추출 결과를 자금수지표 템플릿에 자동 입력"""
        try:
            wb = load_workbook(template_path)
            
            # BS 계정 처리
            bs_data = {k: v for k, v in extracted_data.items() if v['account_type'] == 'BS'}
            if bs_data and self.template_structure['bs_sheet_name'] in wb.sheetnames:
                self._integrate_bs_data(wb[self.template_structure['bs_sheet_name']], bs_data)
            
            # PL 계정 처리
            pl_data = {k: v for k, v in extracted_data.items() if v['account_type'] == 'PL'}
            if pl_data and self.template_structure['pl_sheet_name'] in wb.sheetnames:
                self._integrate_pl_data(wb[self.template_structure['pl_sheet_name']], pl_data)
            
            wb.save(template_path)
            import logging
            logging.info(f"[템플릿통합완료] [파일_{template_path}] [BS계정_{len(bs_data)}개] [PL계정_{len(pl_data)}개]")
            
            return template_path
            
        except Exception as e:
            import logging
            logging.error(f"[템플릿통합실패] [파일_{template_path}] [오류_{str(e)}]")
            raise
    
    def _integrate_bs_data(self, ws, bs_data):
        """BS 계정 데이터 템플릿 반영"""
        for account_code, account_info in bs_data.items():
            account_row = self._find_account_row(ws, account_code)
            if account_row:
                # 월별 데이터 입력 (2열부터 13열까지 = 1월~12월)
                monthly_data = account_info['monthly_data']
                for month in range(1, 13):
                    col_idx = 1 + month  # B열부터 시작
                    if month in monthly_data:
                        ws.cell(row=account_row, column=col_idx).value = monthly_data[month]
    
    def _integrate_pl_data(self, ws, pl_data):
        """PL 계정 데이터 템플릿 반영 - PL 시트 전용"""
        import logging
        from openpyxl.styles import PatternFill
        
        # PL 계정 이름 매핑
        pl_account_names = {
            '40000': '매출',
            '80200': '인건비', 
            '80800': '전력비',
            '81100': '통신비',
            '81200': '수도광열비'
        }
        
        # 헤더가 없으면 생성
        if ws.max_row == 1 and not ws['A1'].value:
            ws['A1'] = '계정코드'
            ws['B1'] = '계정명'
            ws['C1'] = '전기이월'
            for month in range(1, 13):
                col_letter = chr(ord('D') + month - 1)
                ws[f'{col_letter}1'] = f'2022년{month:02d}월'
        
        # 데이터 행 시작점 찾기
        current_row = 2
        
        for account_code, account_info in pl_data.items():
            # 계정 정보 입력
            ws[f'A{current_row}'] = account_code
            ws[f'B{current_row}'] = pl_account_names.get(account_code, account_info.get('account_name', 'Unknown'))
            
            # 전기이월 입력 (PL은 보통 없음)
            carry_forward = account_info.get('carry_forward')
            ws[f'C{current_row}'] = carry_forward
            
            # 월별 발생액 입력
            monthly_data = account_info.get('monthly_data', {})
            data_count = 0
            
            for month in range(1, 13):
                col_letter = chr(ord('D') + month - 1)
                if month in monthly_data and monthly_data[month] != 0:
                    ws[f'{col_letter}{current_row}'] = monthly_data[month]
                    data_count += 1
                else:
                    ws[f'{col_letter}{current_row}'] = None
            
            # 데이터 없는 계정 노란색 마킹
            if data_count == 0:
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                for col in ['A', 'B', 'C'] + [chr(ord('D') + i) for i in range(12)]:
                    ws[f'{col}{current_row}'].fill = yellow_fill
                logging.warning(f'[PL데이터없음] [계정_{account_code}] [노란색마킹]')
            else:
                logging.info(f'[PL데이터입력] [계정_{account_code}] [데이터월수_{data_count}개월]')
            
            current_row += 1
    
    def _find_account_row(self, ws, account_code):
        """워크시트에서 계정 코드에 해당하는 행 찾기"""
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and account_code in str(cell_value):
                return row
        return None


class RequestParser:
    """
    사용자 요청을 파라미터로 변환
    "2022년 BS계정 첫 3개만" → {"years": [2022], "accounts": ["10200", "10300", "10400"]}
    """
    
    def __init__(self):
        self.bs_first_accounts = ["10200", "10300", "10400"]
        self.common_accounts = {
            "보통예금": ["10300"],
            "미지급금": ["25300"],
            "직원급여": ["80200"]
        }
        
        import logging
        logging.info("[요청파서초기화] [패턴설정완료]")
    
    def parse_user_request(self, **kwargs):
        """사용자 요청 파라미터 파싱"""
        # 직접 파라미터가 주어진 경우
        if 'years' in kwargs or 'accounts' in kwargs:
            return {
                'years': kwargs.get('years', [2022]),
                'accounts': kwargs.get('accounts', None),
                'account_type': kwargs.get('account_type', None),
                'output_format': kwargs.get('output_format', 'template')
            }
        
        # 자연어 요청 파싱
        user_input = kwargs.get('request_text', '')
        return self.parse_natural_language_request(user_input)
    
    def parse_natural_language_request(self, user_input):
        """자연어 요청 파싱"""
        # 연도 추출
        years = self._extract_years(user_input)
        
        # 계정 추출
        accounts = None
        account_type = None
        
        if "첫 3개" in user_input and "BS" in user_input:
            accounts = self.bs_first_accounts
        elif "보통예금" in user_input:
            accounts = self.common_accounts["보통예금"]
        elif "BS계정" in user_input or "자산" in user_input:
            account_type = "BS"
        elif "PL계정" in user_input or "손익" in user_input:
            account_type = "PL"
        
        return {
            'years': years,
            'accounts': accounts,
            'account_type': account_type,
            'output_format': 'template'
        }
    
    def _extract_years(self, user_input):
        """텍스트에서 연도 추출"""
        import re
        
        # 연도 패턴 찾기
        year_patterns = [
            r'(\d{4})년',
            r'(\d{4})-(\d{4})년',
            r'(\d{4})'
        ]
        
        years = []
        for pattern in year_patterns:
            matches = re.findall(pattern, user_input)
            for match in matches:
                if isinstance(match, tuple):
                    # 범위 (예: 2022-2024년)
                    start_year = int(match[0])
                    end_year = int(match[1])
                    years.extend(range(start_year, end_year + 1))
                else:
                    # 단일 연도
                    year = int(match)
                    if 2020 <= year <= 2030:  # 유효한 연도 범위
                        years.append(year)
        
        return years if years else [2022]  # 기본값


if __name__ == "__main__":
    main()
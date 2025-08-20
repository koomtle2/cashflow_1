"""
재무데이터 처리 시스템 - 노란색 마킹 시스템
CLAUDE.md 규칙: 모든 불확실한 데이터는 노란색 마킹 후 빈 값 처리
"""

import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from typing import Dict, List, Optional, Any

class YellowMarkingSystem:
    """
    노란색 마킹 시스템
    핵심 원칙: 불확실한 데이터는 노란색 마킹 + 값 비움 + 상세 로깅
    """
    
    def __init__(self):
        # 노란색 마킹 스타일
        self.yellow_fill = PatternFill(
            start_color="FFFF00", 
            end_color="FFFF00", 
            fill_type="solid"
        )
        
        # 텍스트 스타일 (가독성 향상)
        self.marked_font = Font(
            bold=True,
            color="000000"  # 검은색 텍스트
        )
        
        # 정렬
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        
        self.marked_cells = []
        self.marking_stats = {
            'total_marked': 0,
            'by_issue_type': {},
            'by_account': {},
            'by_sheet': {}
        }
        
        logging.info("[노란색마킹시스템초기화] [스타일설정완료]")
    
    def mark_uncertain_cell(self, workbook, sheet_name: str, cell_coordinate: str, 
                          account_code: str, issue_type: str, detail: str, 
                          original_value: Any = None) -> Dict:
        """
        불확실한 데이터 노란색 마킹
        
        Args:
            workbook: 엑셀 워크북 객체
            sheet_name: 시트명
            cell_coordinate: 셀 좌표 (예: 'G5')
            account_code: 계정 코드
            issue_type: 이슈 유형
            detail: 상세 내용
            original_value: 원본 값
        """
        try:
            sheet = workbook[sheet_name]
            cell = sheet[cell_coordinate]
            
            # 원본 값 보존
            if original_value is None:
                original_value = cell.value
            
            # 노란색 마킹 적용
            cell.fill = self.yellow_fill
            cell.font = self.marked_font
            cell.alignment = self.center_alignment
            
            # 값 비우기 (CLAUDE.md 규칙: 추정값 절대 금지)
            cell.value = None
            
            # 코멘트 추가 (사용자 확인용)
            comment_text = f"데이터 이슈: {issue_type}\n상세: {detail}\n확인 필요"
            cell.comment = Comment(comment_text, "시스템")
            
            # 마킹 기록
            marking_record = {
                'timestamp': datetime.now().isoformat(),
                'sheet_name': sheet_name,
                'cell_coordinate': cell_coordinate,
                'account_code': account_code,
                'issue_type': issue_type,
                'detail': detail,
                'original_value': original_value,
                'action': '노란색_마킹_값_비움'
            }
            
            self.marked_cells.append(marking_record)
            
            # 통계 업데이트
            self._update_stats(sheet_name, account_code, issue_type)
            
            # 로깅
            logging.warning(
                f"[노란색마킹적용] [계정_{account_code}] [시트_{sheet_name}] "
                f"[셀_{cell_coordinate}] [이슈_{issue_type}] [상세_{detail}] "
                f"[원본값_{original_value}] [처리_값비움]"
            )
            
            return {
                'marked': True,
                'cell_coordinate': cell_coordinate,
                'original_value': original_value,
                'issue_type': issue_type,
                'detail': detail
            }
            
        except Exception as e:
            logging.error(
                f"[노란색마킹오류] [계정_{account_code}] [시트_{sheet_name}] "
                f"[셀_{cell_coordinate}] [오류_{str(e)}]"
            )
            return {
                'marked': False,
                'error': str(e)
            }
    
    def mark_range_uncertain(self, workbook, sheet_name: str, start_cell: str, 
                           end_cell: str, account_code: str, issue_type: str, 
                           detail: str) -> List[Dict]:
        """범위 셀 노란색 마킹"""
        marked_results = []
        
        try:
            sheet = workbook[sheet_name]
            cell_range = sheet[f"{start_cell}:{end_cell}"]
            
            if isinstance(cell_range, tuple):
                # 2D 범위인 경우
                for row in cell_range:
                    for cell in row:
                        result = self.mark_uncertain_cell(
                            workbook, sheet_name, cell.coordinate,
                            account_code, issue_type, 
                            f"{detail}_범위마킹_{cell.coordinate}"
                        )
                        marked_results.append(result)
            else:
                # 1D 범위인 경우
                for cell in cell_range:
                    result = self.mark_uncertain_cell(
                        workbook, sheet_name, cell.coordinate,
                        account_code, issue_type,
                        f"{detail}_범위마킹_{cell.coordinate}"
                    )
                    marked_results.append(result)
            
            logging.info(
                f"[범위마킹완료] [계정_{account_code}] [시트_{sheet_name}] "
                f"[범위_{start_cell}:{end_cell}] [셀수_{len(marked_results)}개]"
            )
            
        except Exception as e:
            logging.error(
                f"[범위마킹오류] [계정_{account_code}] [시트_{sheet_name}] "
                f"[범위_{start_cell}:{end_cell}] [오류_{str(e)}]"
            )
        
        return marked_results
    
    def mark_contamination_alert(self, workbook, sheet_name: str, 
                               contamination_data: Dict) -> Dict:
        """교차 오염 감지시 마킹"""
        account = contamination_data.get('account', 'UNKNOWN')
        year = contamination_data.get('year', 'UNKNOWN')
        month = contamination_data.get('month', 'UNKNOWN')
        suspicion = contamination_data.get('suspicion', 'UNKNOWN')
        
        # 시트 전체를 위험 상태로 마킹
        try:
            sheet = workbook[sheet_name]
            
            # 첫 번째 행에 경고 메시지 추가
            warning_cell = sheet['A1']
            warning_cell.value = f"⚠️ 교차오염위험: {suspicion}"
            warning_cell.fill = PatternFill(
                start_color="FF0000",  # 빨간색
                end_color="FF0000",
                fill_type="solid"
            )
            warning_cell.font = Font(bold=True, color="FFFFFF")
            
            # 관련 데이터 영역 노란색 마킹
            if month != 'UNKNOWN':
                # 월별 데이터 추정 위치 마킹
                for row in range(1, min(sheet.max_row + 1, 100)):
                    self.mark_uncertain_cell(
                        workbook, sheet_name, f'G{row}',
                        account, "교차오염위험",
                        f"년_{year}_월_{month}_{suspicion}"
                    )
            
            logging.error(
                f"[교차오염마킹] [계정_{account}] [시트_{sheet_name}] "
                f"[년_{year}] [월_{month}] [의심_{suspicion}]"
            )
            
            return {
                'contamination_marked': True,
                'account': account,
                'year': year,
                'month': month,
                'suspicion': suspicion
            }
            
        except Exception as e:
            logging.error(
                f"[교차오염마킹오류] [계정_{account}] [시트_{sheet_name}] [오류_{str(e)}]"
            )
            return {'contamination_marked': False, 'error': str(e)}
    
    def _update_stats(self, sheet_name: str, account_code: str, issue_type: str):
        """마킹 통계 업데이트"""
        self.marking_stats['total_marked'] += 1
        
        # 이슈 유형별 통계
        if issue_type in self.marking_stats['by_issue_type']:
            self.marking_stats['by_issue_type'][issue_type] += 1
        else:
            self.marking_stats['by_issue_type'][issue_type] = 1
        
        # 계정별 통계
        if account_code in self.marking_stats['by_account']:
            self.marking_stats['by_account'][account_code] += 1
        else:
            self.marking_stats['by_account'][account_code] = 1
        
        # 시트별 통계
        if sheet_name in self.marking_stats['by_sheet']:
            self.marking_stats['by_sheet'][sheet_name] += 1
        else:
            self.marking_stats['by_sheet'][sheet_name] = 1
    
    def create_marking_summary_sheet(self, workbook) -> Dict:
        """마킹 요약 시트 생성"""
        try:
            # 기존 요약 시트 삭제 (있는 경우)
            if "마킹요약" in workbook.sheetnames:
                del workbook["마킹요약"]
            
            # 새 요약 시트 생성
            summary_sheet = workbook.create_sheet("마킹요약")
            
            # 헤더 작성
            headers = [
                "번호", "시트명", "셀위치", "계정코드", "이슈유형", 
                "상세내용", "원본값", "마킹시간", "상태"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col)
                cell.value = header
                cell.fill = PatternFill(
                    start_color="CCCCCC",
                    end_color="CCCCCC",
                    fill_type="solid"
                )
                cell.font = Font(bold=True)
            
            # 데이터 작성
            for idx, record in enumerate(self.marked_cells, 2):
                summary_sheet.cell(row=idx, column=1).value = idx - 1
                summary_sheet.cell(row=idx, column=2).value = record['sheet_name']
                summary_sheet.cell(row=idx, column=3).value = record['cell_coordinate']
                summary_sheet.cell(row=idx, column=4).value = record['account_code']
                summary_sheet.cell(row=idx, column=5).value = record['issue_type']
                summary_sheet.cell(row=idx, column=6).value = record['detail']
                summary_sheet.cell(row=idx, column=7).value = record['original_value']
                summary_sheet.cell(row=idx, column=8).value = record['timestamp']
                summary_sheet.cell(row=idx, column=9).value = "확인필요"
                
                # 행 색상 (연한 노란색)
                for col in range(1, 10):
                    summary_sheet.cell(row=idx, column=col).fill = PatternFill(
                        start_color="FFFACD",
                        end_color="FFFACD",
                        fill_type="solid"
                    )
            
            # 통계 섹션 추가
            stats_row = len(self.marked_cells) + 5
            
            summary_sheet.cell(row=stats_row, column=1).value = "=== 마킹 통계 ==="
            summary_sheet.cell(row=stats_row, column=1).font = Font(bold=True, size=14)
            
            summary_sheet.cell(row=stats_row + 1, column=1).value = "총 마킹 개수:"
            summary_sheet.cell(row=stats_row + 1, column=2).value = self.marking_stats['total_marked']
            
            summary_sheet.cell(row=stats_row + 2, column=1).value = "이슈 유형별:"
            for i, (issue_type, count) in enumerate(self.marking_stats['by_issue_type'].items()):
                summary_sheet.cell(row=stats_row + 3 + i, column=1).value = f"  - {issue_type}:"
                summary_sheet.cell(row=stats_row + 3 + i, column=2).value = f"{count}건"
            
            # 열 너비 조정
            column_widths = [8, 20, 12, 15, 20, 30, 15, 20, 12]
            for col, width in enumerate(column_widths, 1):
                summary_sheet.column_dimensions[
                    summary_sheet.cell(row=1, column=col).column_letter
                ].width = width
            
            logging.info(
                f"[마킹요약시트생성] [총마킹_{self.marking_stats['total_marked']}건] "
                f"[이슈유형_{len(self.marking_stats['by_issue_type'])}개]"
            )
            
            return {
                'summary_created': True,
                'total_marked': self.marking_stats['total_marked'],
                'sheet_name': '마킹요약'
            }
            
        except Exception as e:
            logging.error(f"[마킹요약시트오류] [오류_{str(e)}]")
            return {'summary_created': False, 'error': str(e)}
    
    def get_marking_summary(self) -> Dict:
        """마킹 결과 요약 반환"""
        return {
            'timestamp': datetime.now().isoformat(),
            'total_marked_cells': len(self.marked_cells),
            'marking_statistics': self.marking_stats,
            'marked_cells_details': self.marked_cells,
            'data_integrity_status': 'COMPROMISED' if self.marked_cells else 'INTACT'
        }
    
    def export_marking_log(self, file_path: str = None) -> str:
        """마킹 로그를 별도 파일로 내보내기"""
        if not file_path:
            file_path = f"./log/marking_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("=== 노란색 마킹 상세 로그 ===\n")
                f.write(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"총 마킹 개수: {len(self.marked_cells)}건\n\n")
                
                f.write("=== 마킹 통계 ===\n")
                f.write(f"전체 마킹: {self.marking_stats['total_marked']}건\n")
                
                f.write("\n이슈 유형별 분포:\n")
                for issue_type, count in self.marking_stats['by_issue_type'].items():
                    f.write(f"  - {issue_type}: {count}건\n")
                
                f.write("\n계정별 분포:\n")
                for account, count in self.marking_stats['by_account'].items():
                    f.write(f"  - 계정 {account}: {count}건\n")
                
                f.write("\n시트별 분포:\n")
                for sheet, count in self.marking_stats['by_sheet'].items():
                    f.write(f"  - {sheet}: {count}건\n")
                
                f.write("\n=== 상세 마킹 기록 ===\n")
                for i, record in enumerate(self.marked_cells, 1):
                    f.write(f"\n{i}. 시트: {record['sheet_name']}\n")
                    f.write(f"   셀: {record['cell_coordinate']}\n")
                    f.write(f"   계정: {record['account_code']}\n")
                    f.write(f"   이슈: {record['issue_type']}\n")
                    f.write(f"   상세: {record['detail']}\n")
                    f.write(f"   원본값: {record['original_value']}\n")
                    f.write(f"   시간: {record['timestamp']}\n")
            
            logging.info(f"[마킹로그내보내기완료] [파일_{file_path}] [UTF8인코딩]")
            return file_path
            
        except Exception as e:
            logging.error(f"[마킹로그내보내기오류] [파일_{file_path}] [오류_{str(e)}]")
            return ""
    
    def validate_marking_integrity(self) -> Dict:
        """마킹 무결성 검증"""
        validation_result = {
            'integrity_check_passed': True,
            'issues_found': [],
            'recommendations': []
        }
        
        # 중복 마킹 검사
        cell_coordinates = [record['cell_coordinate'] + '_' + record['sheet_name'] 
                           for record in self.marked_cells]
        duplicates = [coord for coord in set(cell_coordinates) 
                     if cell_coordinates.count(coord) > 1]
        
        if duplicates:
            validation_result['integrity_check_passed'] = False
            validation_result['issues_found'].append(f"중복_마킹_{len(duplicates)}건")
            validation_result['recommendations'].append("중복_마킹_제거_필요")
        
        # 마킹 통계 일관성 검사
        calculated_total = sum(self.marking_stats['by_issue_type'].values())
        if calculated_total != self.marking_stats['total_marked']:
            validation_result['integrity_check_passed'] = False
            validation_result['issues_found'].append("통계_불일치")
            validation_result['recommendations'].append("통계_재계산_필요")
        
        # 빈 마킹 레코드 검사
        empty_records = [record for record in self.marked_cells 
                        if not record.get('account_code') or not record.get('issue_type')]
        
        if empty_records:
            validation_result['integrity_check_passed'] = False
            validation_result['issues_found'].append(f"빈레코드_{len(empty_records)}건")
            validation_result['recommendations'].append("빈레코드_정리_필요")
        
        if validation_result['integrity_check_passed']:
            logging.info("[마킹무결성검증통과] [이상없음]")
        else:
            logging.warning(
                f"[마킹무결성검증실패] [이슈_{len(validation_result['issues_found'])}개] "
                f"[권고사항_{len(validation_result['recommendations'])}개]"
            )
        
        return validation_result
    
    def detect_cross_contamination(self, v3_data: Dict, ledger_data: Dict) -> List[Dict]:
        """교차 오염 실시간 감지 시스템 (CLAUDE.md 17. 규칙)"""
        contamination_alerts = []
        
        logging.info("[교차오염감지시작] [v3데이터_vs_원장데이터_비교]")
        
        for account_code in v3_data.keys():
            for year, year_data in v3_data[account_code].items():
                for month, v3_value in year_data.items():
                    ledger_value = ledger_data.get(year, {}).get(account_code, {}).get(month, 0)
                    
                    # 패턴 1: v3에 외부 데이터 유입 의심
                    if v3_value != 0 and ledger_value == 0:
                        alert = {
                            'account': account_code,
                            'year': year,
                            'month': month,
                            'v3_value': v3_value,
                            'ledger_value': ledger_value,
                            'suspicion': 'v3에_외부데이터_유입_의심',
                            'contamination_type': 'EXTERNAL_DATA_INJECTION',
                            'risk_level': 'HIGH'
                        }
                        contamination_alerts.append(alert)
                        logging.error(
                            f"[교차오염감지] [유형_외부데이터유입] [계정_{account_code}] "
                            f"[년_{year}] [월_{month}] [v3값_{v3_value}] [원장값_{ledger_value}]"
                        )
                    
                    # 패턴 2: 부호 반전 오류 감지
                    elif v3_value * ledger_value < 0 and abs(v3_value) == abs(ledger_value):
                        alert = {
                            'account': account_code,
                            'year': year,
                            'month': month,
                            'v3_value': v3_value,
                            'ledger_value': ledger_value,
                            'suspicion': '부호반전_처리오류',
                            'contamination_type': 'SIGN_REVERSAL_ERROR',
                            'risk_level': 'MEDIUM'
                        }
                        contamination_alerts.append(alert)
                        logging.warning(
                            f"[교차오염감지] [유형_부호반전] [계정_{account_code}] "
                            f"[년_{year}] [월_{month}] [v3값_{v3_value}] [원장값_{ledger_value}]"
                        )
                    
                    # 패턴 3: 수익 계정 음수값 감지
                    elif self._is_revenue_account(account_code) and v3_value < 0:
                        alert = {
                            'account': account_code,
                            'year': year,
                            'month': month,
                            'v3_value': v3_value,
                            'ledger_value': ledger_value,
                            'suspicion': '수익계정_음수값_오염',
                            'contamination_type': 'REVENUE_NEGATIVE_VALUE',
                            'risk_level': 'HIGH'
                        }
                        contamination_alerts.append(alert)
                        logging.error(
                            f"[교차오염감지] [유형_수익계정음수] [계정_{account_code}] "
                            f"[년_{year}] [월_{month}] [v3값_{v3_value}] [수익계정_음수불가]"
                        )
        
        # 패턴 4: 동일 금액 여러 계정 중복 감지
        duplicate_amounts = self._detect_duplicate_amounts(v3_data)
        for duplicate_info in duplicate_amounts:
            alert = {
                'accounts': duplicate_info['accounts'],
                'year': duplicate_info['year'],
                'month': duplicate_info['month'],
                'duplicate_value': duplicate_info['value'],
                'suspicion': '동일금액_중복_교차오염',
                'contamination_type': 'DUPLICATE_AMOUNT_CROSS_CONTAMINATION',
                'risk_level': 'HIGH'
            }
            contamination_alerts.append(alert)
            logging.error(
                f"[교차오염감지] [유형_중복금액] [계정들_{duplicate_info['accounts']}] "
                f"[년_{duplicate_info['year']}] [월_{duplicate_info['month']}] "
                f"[중복값_{duplicate_info['value']}]"
            )
        
        if contamination_alerts:
            logging.error(
                f"[교차오염감지완료] [총발견_{len(contamination_alerts)}건] "
                f"[HIGH위험_{len([a for a in contamination_alerts if a.get('risk_level') == 'HIGH'])}건]"
            )
        else:
            logging.info("[교차오염감지완료] [이상없음] [데이터무결성유지]")
        
        return contamination_alerts
    
    def _is_revenue_account(self, account_code: str) -> bool:
        """수익 계정 여부 판단"""
        try:
            code_num = int(account_code)
            # 수익 계정 범위: 40000-42100, 90000-92100
            revenue_ranges = [(40000, 42100), (90000, 92100)]
            return any(start <= code_num <= end for start, end in revenue_ranges)
        except (ValueError, TypeError):
            return False
    
    def _detect_duplicate_amounts(self, v3_data: Dict) -> List[Dict]:
        """동일 금액 중복 감지"""
        duplicate_amounts = []
        amount_map = {}  # {(year, month, value): [account_codes]}
        
        for account_code, account_data in v3_data.items():
            for year, year_data in account_data.items():
                for month, value in year_data.items():
                    if value != 0:  # 0값은 제외
                        key = (year, month, value)
                        if key not in amount_map:
                            amount_map[key] = []
                        amount_map[key].append(account_code)
        
        # 중복 발견
        for (year, month, value), accounts in amount_map.items():
            if len(accounts) > 1:
                duplicate_amounts.append({
                    'year': year,
                    'month': month,
                    'value': value,
                    'accounts': accounts
                })
        
        return duplicate_amounts
    
    def mark_all_contamination_alerts(self, workbook, contamination_alerts: List[Dict]) -> Dict:
        """모든 교차 오염 알람을 시각적으로 마킹"""
        marking_summary = {
            'total_alerts': len(contamination_alerts),
            'marked_sheets': [],
            'high_risk_count': 0,
            'medium_risk_count': 0
        }
        
        for alert in contamination_alerts:
            # 위험도별 카운트
            if alert.get('risk_level') == 'HIGH':
                marking_summary['high_risk_count'] += 1
            elif alert.get('risk_level') == 'MEDIUM':
                marking_summary['medium_risk_count'] += 1
            
            # 관련 시트에 마킹
            if 'account' in alert:
                # 단일 계정 오염
                sheet_name = f"오염감지_{alert['account']}"
                if sheet_name not in workbook.sheetnames:
                    # 새 시트 생성
                    warning_sheet = workbook.create_sheet(sheet_name)
                    self._create_contamination_warning_sheet(warning_sheet, alert)
                    marking_summary['marked_sheets'].append(sheet_name)
            
            elif 'accounts' in alert:
                # 중복 금액 오염
                sheet_name = "오염감지_중복금액"
                if sheet_name not in workbook.sheetnames:
                    warning_sheet = workbook.create_sheet(sheet_name)
                    self._create_duplicate_warning_sheet(warning_sheet, alert)
                    marking_summary['marked_sheets'].append(sheet_name)
        
        logging.info(
            f"[교차오염마킹완료] [총알람_{marking_summary['total_alerts']}건] "
            f"[HIGH위험_{marking_summary['high_risk_count']}건] "
            f"[마킹시트_{len(marking_summary['marked_sheets'])}개]"
        )
        
        return marking_summary
    
    def _create_contamination_warning_sheet(self, sheet, alert: Dict):
        """교차 오염 경고 시트 생성"""
        # 헤더
        sheet['A1'] = "⚠️ 교차 오염 감지 알람"
        sheet['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        
        # 상세 정보
        sheet['A3'] = "계정 코드:"
        sheet['B3'] = alert['account']
        sheet['A4'] = "연도:"
        sheet['B4'] = alert['year']
        sheet['A5'] = "월:"
        sheet['B5'] = alert['month']
        sheet['A6'] = "의심 유형:"
        sheet['B6'] = alert['suspicion']
        sheet['A7'] = "위험도:"
        sheet['B7'] = alert['risk_level']
        sheet['A8'] = "v3 값:"
        sheet['B8'] = alert['v3_value']
        sheet['A9'] = "원장 값:"
        sheet['B9'] = alert['ledger_value']
        
        # 경고 색상 적용
        for row in range(3, 10):
            sheet[f'A{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def _create_duplicate_warning_sheet(self, sheet, alert: Dict):
        """중복 금액 경고 시트 생성"""
        # 헤더
        sheet['A1'] = "⚠️ 중복 금액 교차 오염 감지"
        sheet['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        sheet['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        
        # 상세 정보
        sheet['A3'] = "관련 계정들:"
        sheet['B3'] = ", ".join(alert['accounts'])
        sheet['A4'] = "연도:"
        sheet['B4'] = alert['year']
        sheet['A5'] = "월:"
        sheet['B5'] = alert['month']
        sheet['A6'] = "중복 금액:"
        sheet['B6'] = alert['duplicate_value']
        sheet['A7'] = "의심 유형:"
        sheet['B7'] = alert['suspicion']
        
        # 경고 색상 적용
        for row in range(3, 8):
            sheet[f'A{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet[f'B{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def real_time_contamination_monitor(self, account_code: str, new_data: Dict, 
                                      existing_data: Dict) -> bool:
        """실시간 교차 오염 모니터링 (데이터 쓰기 전 사전 차단)"""
        
        # 1. 수익 계정 음수값 즉시 차단
        if self._is_revenue_account(account_code):
            for month, value in new_data.items():
                if value < 0:
                    logging.error(
                        f"[실시간오염차단] [수익계정음수] [계정_{account_code}] "
                        f"[월_{month}] [값_{value}] [처리중단]"
                    )
                    return False
        
        # 2. 동일 금액 중복 즉시 감지
        for month, new_value in new_data.items():
            if new_value != 0:
                for existing_account, existing_months in existing_data.items():
                    if existing_account != account_code and month in existing_months:
                        if existing_months[month] == new_value:
                            logging.error(
                                f"[실시간오염차단] [중복금액] [계정_{account_code}] "
                                f"[기존계정_{existing_account}] [월_{month}] [값_{new_value}] [처리중단]"
                            )
                            return False
        
        # 3. 원장 없는 계정에 데이터 입력 시도 차단
        if account_code not in existing_data and any(v != 0 for v in new_data.values()):
            logging.warning(
                f"[실시간오염경고] [원장없는계정] [계정_{account_code}] "
                f"[데이터입력시도] [확인필요]"
            )
            # 경고만 하고 차단하지는 않음 (새로운 계정일 가능성)
        
        logging.info(
            f"[실시간모니터링통과] [계정_{account_code}] [정상처리가능]"
        )
        return True
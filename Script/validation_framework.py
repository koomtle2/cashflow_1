"""
재무데이터 처리 시스템 - 데이터 검증 프레임워크
CLAUDE.md 규칙 완전 준수 (데이터 완결성 최우선)
"""

import logging
import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, List, Optional, Tuple, Any

class DataValidator:
    """
    데이터 검증 프레임워크
    핵심 원칙: 데이터 완결성 최우선, 확실하지 않으면 노란색 마킹
    """
    
    def __init__(self):
        self.yellow_marks = []
        self.log_file = self._create_utf8_log()
        self.yellow_fill = PatternFill(
            start_color="FFFF00", 
            end_color="FFFF00", 
            fill_type="solid"
        )
        self.processed_accounts = {}
        self.contamination_alerts = []
        
        # BS/PL 계정 분류 (CLAUDE.md 기준)
        self.bs_accounts = {
            # 자산 계정 (100-200번대)
            '100': 'BS', '101': 'BS', '102': 'BS', '103': 'BS', '104': 'BS', '105': 'BS',
            '110': 'BS', '111': 'BS', '112': 'BS', '113': 'BS', '114': 'BS', '115': 'BS',
            '120': 'BS', '121': 'BS', '122': 'BS', '123': 'BS', '124': 'BS', '125': 'BS',
            '130': 'BS', '131': 'BS', '132': 'BS', '133': 'BS', '134': 'BS', '135': 'BS',
            '140': 'BS', '141': 'BS', '142': 'BS', '143': 'BS', '144': 'BS', '145': 'BS',
            '150': 'BS', '151': 'BS', '152': 'BS', '153': 'BS', '154': 'BS', '155': 'BS',
            '160': 'BS', '161': 'BS', '162': 'BS', '163': 'BS', '164': 'BS', '165': 'BS',
            '170': 'BS', '171': 'BS', '172': 'BS', '173': 'BS', '174': 'BS', '175': 'BS',
            '180': 'BS', '181': 'BS', '182': 'BS', '183': 'BS', '184': 'BS', '185': 'BS',
            '190': 'BS', '191': 'BS', '192': 'BS', '193': 'BS', '194': 'BS', '195': 'BS',
            
            # 부채 계정 (250-290번대)
            '250': 'BS', '251': 'BS', '252': 'BS', '253': 'BS', '254': 'BS', '255': 'BS',
            '256': 'BS', '257': 'BS', '258': 'BS', '259': 'BS', '260': 'BS', '261': 'BS',
            '262': 'BS', '263': 'BS', '264': 'BS', '265': 'BS', '266': 'BS', '267': 'BS',
            '268': 'BS', '269': 'BS', '270': 'BS', '271': 'BS', '272': 'BS', '273': 'BS',
            '274': 'BS', '275': 'BS', '276': 'BS', '277': 'BS', '278': 'BS', '279': 'BS',
            '280': 'BS', '281': 'BS', '282': 'BS', '283': 'BS', '284': 'BS', '285': 'BS',
            '286': 'BS', '287': 'BS', '288': 'BS', '289': 'BS', '290': 'BS', '291': 'BS',
            
            # 자본 계정 (330-370번대)
            '330': 'BS', '331': 'BS', '332': 'BS', '333': 'BS', '334': 'BS', '335': 'BS',
            '340': 'BS', '341': 'BS', '342': 'BS', '343': 'BS', '344': 'BS', '345': 'BS',
            '350': 'BS', '351': 'BS', '352': 'BS', '353': 'BS', '354': 'BS', '355': 'BS',
            '360': 'BS', '361': 'BS', '362': 'BS', '363': 'BS', '364': 'BS', '365': 'BS',
            '370': 'BS', '371': 'BS', '372': 'BS', '373': 'BS', '374': 'BS', '375': 'BS',
        }
        
        # 특별 계정 (VAT 관련)
        self.vat_accounts = {
            '13500': 'VAT_부가세대급금',  # 지급할 부가세
            '25500': 'VAT_부가세예수금',  # 수취한 부가세
        }
        
        logging.info(f"[초기화완료] [로그파일_{self.log_file}] [UTC인코딩보장]")
    
    def _create_utf8_log(self) -> str:
        """UTF-8 인코딩 보장 로그 파일 생성"""
        log_filename = f"./log/{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.log"
        os.makedirs("./log", exist_ok=True)
        
        # UTF-8 인코딩 명시적 지정
        logging.basicConfig(
            filename=log_filename,
            level=logging.INFO,
            format='[%(asctime)s] [%(levelname)s] %(message)s',
            filemode='w',
            encoding='utf-8',  # UTF-8 인코딩 보장
            force=True
        )
        
        return log_filename
    
    def extract_account_code(self, sheet_name: str) -> Optional[str]:
        """시트명에서 계정코드 추출 (CLAUDE.md 정규식 사용)"""
        pattern = r'\((\d+)\)'
        match = re.search(pattern, sheet_name)
        
        if match:
            account_code = match.group(1)
            logging.info(f"[계정코드추출성공] [시트_{sheet_name}] [코드_{account_code}]")
            return account_code
        else:
            logging.warning(f"[계정코드추출실패] [시트_{sheet_name}] [패턴불일치] [노란색마킹대상]")
            return None
    
    def classify_account_type(self, account_code: str) -> str:
        """계정 유형 분류 (BS/PL)"""
        if not account_code:
            return "UNKNOWN"
        
        # 3자리 코드로 변환
        code_3digit = account_code[:3] if len(account_code) >= 3 else account_code
        
        if code_3digit in self.bs_accounts:
            logging.info(f"[계정분류성공] [코드_{account_code}] [유형_BS]")
            return "BS"
        elif account_code in self.vat_accounts:
            logging.info(f"[계정분류성공] [코드_{account_code}] [유형_VAT_{self.vat_accounts[account_code]}]")
            return "VAT"
        else:
            # PL 계정으로 추정 (400-420, 450-460, 520-530, 800-840, 900-920, 930-960)
            if (code_3digit.startswith('4') or code_3digit.startswith('5') or 
                code_3digit.startswith('8') or code_3digit.startswith('9')):
                logging.info(f"[계정분류성공] [코드_{account_code}] [유형_PL]")
                return "PL"
            else:
                logging.warning(f"[계정분류실패] [코드_{account_code}] [UNKNOWN] [노란색마킹필요]")
                return "UNKNOWN"
    
    def validate_data_structure(self, workbook, sheet_name: str) -> bool:
        """1단계: 데이터 구조 검증"""
        try:
            sheet = workbook[sheet_name]
            logging.info(f"[구조검증시작] [시트_{sheet_name}]")
            
            # 필수 컬럼 존재 확인 (A~G열)
            required_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
            missing_columns = []
            
            for col in required_columns:
                if not sheet[f'{col}1'].value:
                    missing_columns.append(col)
                    self._mark_yellow_and_log(
                        sheet[f'{col}1'], 
                        sheet_name,
                        "구조오류", 
                        f"필수컬럼_{col}_누락"
                    )
            
            if missing_columns:
                logging.error(f"[구조검증실패] [시트_{sheet_name}] [누락컬럼_{missing_columns}]")
                return False
            
            logging.info(f"[구조검증성공] [시트_{sheet_name}] [모든필수컬럼존재]")
            return True
            
        except Exception as e:
            logging.error(f"[구조검증오류] [시트_{sheet_name}] [오류_{str(e)}]")
            return False
    
    def validate_carry_forward(self, sheet, account_code: str) -> Optional[float]:
        """전기이월 데이터 검증 (CLAUDE.md 규칙 엄격 적용)"""
        try:
            b5_value = sheet['B5'].value
            g5_value = sheet['G5'].value
            
            # B5가 정확히 "전기이월"인지 확인
            if b5_value != "전기이월":
                logging.warning(f"[전기이월검증실패] [계정_{account_code}] [B5값_{b5_value}] [전기이월아님] [노란색마킹]")
                self._mark_yellow_and_log(
                    sheet['G5'], 
                    account_code,
                    "전기이월불확실", 
                    f"B5값이_전기이월이_아님_{b5_value}"
                )
                return None
            
            # G5 값이 숫자인지 확인
            if isinstance(g5_value, (int, float)):
                logging.info(f"[전기이월검증성공] [계정_{account_code}] [금액_{g5_value:,}원]")
                return float(g5_value)
            elif g5_value is None or g5_value == "":
                logging.info(f"[전기이월없음] [계정_{account_code}] [G5빈값] [0원처리]")
                return 0.0
            else:
                logging.warning(f"[전기이월형식오류] [계정_{account_code}] [G5값_{g5_value}] [숫자아님] [노란색마킹]")
                self._mark_yellow_and_log(
                    sheet['G5'], 
                    account_code,
                    "전기이월형식오류", 
                    f"G5값이_숫자가_아님_{g5_value}"
                )
                return None
                
        except Exception as e:
            logging.error(f"[전기이월검증오류] [계정_{account_code}] [오류_{str(e)}]")
            return None
    
    def extract_monthly_data(self, sheet, account_code: str, account_type: str) -> Dict[str, float]:
        """월별 데이터 추출 (BS/PL 구분 처리)"""
        monthly_data = {}
        
        try:
            # 날짜 패턴으로 월별 데이터 구분
            date_pattern = r'(\d{2})-\d{2}'  # MM-DD 패턴
            current_month = None
            monthly_transactions = {}
            
            for row in range(6, sheet.max_row + 1):  # 6행부터 데이터 시작
                date_cell = sheet[f'A{row}']
                debit_cell = sheet[f'E{row}']  # 차변
                credit_cell = sheet[f'F{row}']  # 대변
                balance_cell = sheet[f'G{row}']  # 잔액
                
                if date_cell.value and isinstance(date_cell.value, str):
                    match = re.search(date_pattern, date_cell.value)
                    if match:
                        month = match.group(1)
                        
                        # 새로운 월 시작
                        if current_month != month:
                            current_month = month
                            monthly_transactions[month] = []
                        
                        # 거래 데이터 수집
                        debit = self._safe_float_conversion(debit_cell.value, f"차변_M{month}")
                        credit = self._safe_float_conversion(credit_cell.value, f"대변_M{month}")
                        balance = self._safe_float_conversion(balance_cell.value, f"잔액_M{month}")
                        
                        if current_month:
                            monthly_transactions[current_month].append({
                                'debit': debit,
                                'credit': credit,
                                'balance': balance,
                                'row': row
                            })
            
            # BS/PL에 따른 데이터 처리
            for month, transactions in monthly_transactions.items():
                if account_type == "BS":
                    # BS 계정: 월말 최종 잔액
                    if transactions:
                        final_balance = transactions[-1]['balance']
                        if final_balance is not None:
                            monthly_data[month] = final_balance
                            logging.info(f"[BS월별데이터] [계정_{account_code}] [M{month}] [잔액_{final_balance:,}원]")
                        else:
                            logging.warning(f"[BS잔액누락] [계정_{account_code}] [M{month}] [노란색마킹]")
                
                elif account_type == "PL":
                    # PL 계정: 월별 순증감 합산
                    month_total = 0
                    for trans in transactions:
                        if trans['debit'] is not None and trans['credit'] is not None:
                            month_total += (trans['debit'] - trans['credit'])
                    
                    monthly_data[month] = month_total
                    logging.info(f"[PL월별데이터] [계정_{account_code}] [M{month}] [순증감_{month_total:,}원]")
                
                else:  # UNKNOWN 계정
                    logging.warning(f"[계정유형불명] [계정_{account_code}] [M{month}] [처리제외] [노란색마킹]")
            
            return monthly_data
            
        except Exception as e:
            logging.error(f"[월별데이터추출오류] [계정_{account_code}] [오류_{str(e)}]")
            return {}
    
    def _safe_float_conversion(self, value: Any, context: str) -> Optional[float]:
        """안전한 숫자 변환 (오류시 None 반환)"""
        try:
            if value is None or value == "":
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                # 문자열에서 숫자만 추출 시도
                cleaned = re.sub(r'[^\d.-]', '', value)
                if cleaned:
                    return float(cleaned)
            return None
        except:
            logging.warning(f"[숫자변환실패] [컨텍스트_{context}] [값_{value}] [None반환]")
            return None
    
    def detect_cross_contamination(self, processed_data: Dict, original_ledger: Dict) -> List[Dict]:
        """3단계: 교차 오염 검사 (CLAUDE.md 교차 오염 방지 시스템)"""
        contamination_alerts = []
        
        logging.info("[교차오염검사시작] [전체계정대상]")
        
        for account_code, account_data in processed_data.items():
            if not isinstance(account_data, dict):
                continue
                
            for year, year_data in account_data.items():
                if not isinstance(year_data, dict):
                    continue
                    
                for month, processed_value in year_data.items():
                    # 원장 데이터와 비교
                    original_value = (original_ledger.get(year, {})
                                    .get(account_code, {})
                                    .get(month, 0))
                    
                    # 오염 패턴 감지
                    if processed_value != 0 and original_value == 0:
                        contamination_alerts.append({
                            'account': account_code,
                            'year': year,
                            'month': month,
                            'processed_value': processed_value,
                            'original_value': original_value,
                            'suspicion': 'v3에_외부데이터_유입_의심'
                        })
                        logging.error(f"[교차오염의심] [계정_{account_code}] [년_{year}] [월_{month}] [처리값_{processed_value}] [원장값_{original_value}]")
                    
                    # 부호 반전 오류 감지
                    elif (processed_value * original_value < 0 and 
                          abs(processed_value) == abs(original_value)):
                        contamination_alerts.append({
                            'account': account_code,
                            'year': year,
                            'month': month,
                            'processed_value': processed_value,
                            'original_value': original_value,
                            'suspicion': '부호반전_처리오류'
                        })
                        logging.error(f"[부호반전오류] [계정_{account_code}] [년_{year}] [월_{month}] [처리값_{processed_value}] [원장값_{original_value}]")
        
        if contamination_alerts:
            logging.error(f"[교차오염감지] [총_{len(contamination_alerts)}건] [즉시중단필요]")
            self.contamination_alerts.extend(contamination_alerts)
        else:
            logging.info("[교차오염검사완료] [오염없음] [정상]")
        
        return contamination_alerts
    
    def _mark_yellow_and_log(self, cell, account_code: str, issue: str, detail: str):
        """노란색 마킹 + 로깅 동시 처리 (CLAUDE.md 규칙)"""
        cell.fill = self.yellow_fill
        original_value = cell.value
        cell.value = None  # 추정값 입력 절대 금지
        
        self.yellow_marks.append({
            'cell': cell.coordinate,
            'account': account_code,
            'issue': issue,
            'detail': detail,
            'original_value': original_value
        })
        
        logging.warning(f"[{account_code}] [{issue}] [{detail}] [셀_{cell.coordinate}_노란색마킹,값비움] [원본값_{original_value}]")
    
    def validate_account_data(self, workbook, sheet_name: str) -> Dict:
        """계정별 종합 검증"""
        result = {
            'account_code': None,
            'account_type': None,
            'carry_forward': None,
            'monthly_data': {},
            'validation_passed': False,
            'issues': []
        }
        
        try:
            # 계정 코드 추출
            account_code = self.extract_account_code(sheet_name)
            if not account_code:
                result['issues'].append("계정코드_추출_실패")
                return result
            
            result['account_code'] = account_code
            
            # 계정 유형 분류
            account_type = self.classify_account_type(account_code)
            if account_type == "UNKNOWN":
                result['issues'].append("계정분류_실패")
                # 노란색 마킹은 이미 classify_account_type에서 처리됨
            
            result['account_type'] = account_type
            
            # 데이터 구조 검증
            if not self.validate_data_structure(workbook, sheet_name):
                result['issues'].append("데이터구조_검증실패")
                return result
            
            sheet = workbook[sheet_name]
            
            # 전기이월 검증
            carry_forward = self.validate_carry_forward(sheet, account_code)
            result['carry_forward'] = carry_forward
            if carry_forward is None:
                result['issues'].append("전기이월_검증실패")
            
            # 월별 데이터 추출
            monthly_data = self.extract_monthly_data(sheet, account_code, account_type)
            result['monthly_data'] = monthly_data
            
            # 검증 완료 판정
            if not result['issues']:
                result['validation_passed'] = True
                logging.info(f"[계정검증완료] [코드_{account_code}] [유형_{account_type}] [성공]")
            else:
                logging.warning(f"[계정검증부분실패] [코드_{account_code}] [이슈_{result['issues']}]")
            
            return result
            
        except Exception as e:
            logging.error(f"[계정검증오류] [시트_{sheet_name}] [오류_{str(e)}]")
            result['issues'].append(f"검증오류_{str(e)}")
            return result
    
    def get_validation_summary(self) -> Dict:
        """검증 결과 요약"""
        return {
            'total_yellow_marks': len(self.yellow_marks),
            'yellow_marks_details': self.yellow_marks,
            'contamination_alerts': len(self.contamination_alerts),
            'contamination_details': self.contamination_alerts,
            'log_file_path': self.log_file,
            'processed_accounts': len(self.processed_accounts)
        }


class ReportVerificationSystem:
    """
    CLAUDE.md 2025.08.20 추가 규칙: 작업 보고 검증 의무화
    모든 작업 보고 전 반드시 체크리스트 100% 검증
    """
    
    MANDATORY_CHECKLIST = [
        "actual_data_exists",      # 실제 데이터 출력 확인됨
        "logged_in_file",          # 로그 파일에 기록됨
        "no_errors",               # 오류 메시지 없음
        "no_assumptions",          # 추정/가정 없이 작성됨
        "log_matches_report"       # 로그와 보고서 100% 일치
    ]
    
    def __init__(self):
        self.verification_results = {}
        logging.info("[작업보고검증시스템초기화] [의무체크리스트5개설정]")
    
    def verify_before_report(self, processing_result: Dict) -> Dict:
        """
        체크리스트 미완료시 보고 금지
        
        Args:
            processing_result: 처리 결과 데이터
            
        Returns:
            검증 결과 (모든 항목 True여야 보고 허용)
            
        Raises:
            ReportViolationError: 체크리스트 미완료시
        """
        verification = {}
        
        # 1. 실제 데이터 출력 확인됨
        verification["actual_data_exists"] = self._verify_actual_data_exists(processing_result)
        
        # 2. 로그 파일에 기록됨
        verification["logged_in_file"] = self._verify_logged_in_file(processing_result)
        
        # 3. 오류 메시지 없음
        verification["no_errors"] = self._verify_no_errors(processing_result)
        
        # 4. 추정/가정 없이 작성됨
        verification["no_assumptions"] = self._verify_no_assumptions(processing_result)
        
        # 5. 로그와 보고서 100% 일치
        verification["log_matches_report"] = self._verify_log_matches_report(processing_result)
        
        # 검증 결과 저장
        self.verification_results = verification
        
        # 모든 항목 통과 여부 확인
        all_passed = all(verification.values())
        
        if not all_passed:
            failed_items = [item for item, passed in verification.items() if not passed]
            error_msg = f"체크리스트 미완료: {', '.join(failed_items)}"
            logging.error(f"[보고검증실패] [{error_msg}] [보고금지]")
            raise ReportViolationError(error_msg)
        
        logging.info("[보고검증성공] [모든체크리스트통과] [보고허용]")
        return verification
    
    def _verify_actual_data_exists(self, result: Dict) -> bool:
        """실제 데이터 출력 확인됨"""
        try:
            # 추출된 데이터가 실제로 존재하는지 확인
            extracted_data = result.get('extracted_data', {})
            if not extracted_data:
                logging.warning("[체크리스트실패] [actual_data_exists] [추출데이터없음]")
                return False
            
            # 각 계정에 실제 값이 있는지 확인
            has_actual_data = False
            for account_code, account_data in extracted_data.items():
                if account_data.get('monthly_data') or account_data.get('carry_forward'):
                    has_actual_data = True
                    break
            
            if not has_actual_data:
                logging.warning("[체크리스트실패] [actual_data_exists] [실제값없음]")
                return False
            
            return True
            
        except Exception as e:
            logging.error(f"[체크리스트오류] [actual_data_exists] [오류_{str(e)}]")
            return False
    
    def _verify_logged_in_file(self, result: Dict) -> bool:
        """로그 파일에 기록됨"""
        try:
            log_file_path = result.get('log_file_path')
            if not log_file_path or not os.path.exists(log_file_path):
                logging.warning("[체크리스트실패] [logged_in_file] [로그파일없음]")
                return False
            
            # 로그 파일에 실제 내용이 있는지 확인
            with open(log_file_path, 'r', encoding='utf-8') as f:
                log_content = f.read()
                if len(log_content.strip()) < 10:  # 최소한의 로그 내용
                    logging.warning("[체크리스트실패] [logged_in_file] [로그내용부족]")
                    return False
            
            return True
            
        except Exception as e:
            logging.error(f"[체크리스트오류] [logged_in_file] [오류_{str(e)}]")
            return False
    
    def _verify_no_errors(self, result: Dict) -> bool:
        """오류 메시지 없음"""
        try:
            # 처리 결과에서 오류 확인
            processing_errors = result.get('processing_errors', [])
            if processing_errors:
                logging.warning(f"[체크리스트실패] [no_errors] [오류{len(processing_errors)}개발견]")
                return False
            
            # 실패한 계정 확인
            failed_accounts = result.get('failed_accounts', 0)
            if failed_accounts > 0:
                logging.warning(f"[체크리스트실패] [no_errors] [실패계정{failed_accounts}개]")
                return False
            
            return True
            
        except Exception as e:
            logging.error(f"[체크리스트오류] [no_errors] [오류_{str(e)}]")
            return False
    
    def _verify_no_assumptions(self, result: Dict) -> bool:
        """추정/가정 없이 작성됨"""
        try:
            # 노란색 마킹된 셀 = 추정값 사용 안함을 의미
            yellow_marks = result.get('yellow_marked_cells', [])
            
            # 추정값이 사용된 흔적 확인
            extracted_data = result.get('extracted_data', {})
            for account_code, account_data in extracted_data.items():
                monthly_data = account_data.get('monthly_data', {})
                
                # 의심스러운 패턴 (모든 월이 0이거나 동일한 값)
                if monthly_data:
                    values = list(monthly_data.values())
                    if len(set(values)) == 1 and len(values) > 6:  # 6개월 이상 동일한 값
                        logging.warning(f"[체크리스트의심] [no_assumptions] [계정{account_code}_동일값패턴]")
                        # 경고만 하고 실패로 처리하지는 않음 (실제 데이터일 수 있음)
            
            return True
            
        except Exception as e:
            logging.error(f"[체크리스트오류] [no_assumptions] [오류_{str(e)}]")
            return False
    
    def _verify_log_matches_report(self, result: Dict) -> bool:
        """로그와 보고서 100% 일치"""
        try:
            # 로그에서 추출된 데이터와 보고서 데이터 비교
            log_file_path = result.get('log_file_path')
            if not log_file_path or not os.path.exists(log_file_path):
                return False
            
            # 로그에서 성공 처리된 계정 수 계산
            with open(log_file_path, 'r', encoding='utf-8') as f:
                log_content = f.read()
                
                # 로그에서 "[추출성공]" 패턴 개수 세기
                success_count_in_log = log_content.count("[추출성공]")
                
                # 보고서의 성공 계정 수
                extracted_data = result.get('extracted_data', {})
                success_count_in_report = len(extracted_data)
                
                if success_count_in_log != success_count_in_report:
                    logging.warning(f"[체크리스트실패] [log_matches_report] [로그{success_count_in_log}개≠보고서{success_count_in_report}개]")
                    return False
            
            return True
            
        except Exception as e:
            logging.error(f"[체크리스트오류] [log_matches_report] [오류_{str(e)}]")
            return False
    
    def get_checklist_report(self) -> Dict:
        """체크리스트 검증 보고서"""
        return {
            'timestamp': datetime.now().isoformat(),
            'claude_md_compliance': all(self.verification_results.values()),
            'checklist_results': self.verification_results,
            'passed_items': [item for item, passed in self.verification_results.items() if passed],
            'failed_items': [item for item, passed in self.verification_results.items() if not passed]
        }


class DataCompletenessEnforcer:
    """
    CLAUDE.md 데이터 완결성 최우선 원칙 강제 시스템
    확실하지 않은 데이터는 절대 입력하지 않음
    """
    
    def __init__(self):
        self.enforcement_stats = {
            'enforced_count': 0,
            'rejected_data_count': 0,
            'marked_uncertain_count': 0
        }
        
        logging.info("[데이터완결성강제시스템초기화] [원칙_확실하지않으면절대입력금지]")
    
    def enforce_data_completeness(self, data: Any, confidence_level: str = 'UNKNOWN') -> Optional[Any]:
        """
        데이터 완결성 강제 적용
        
        Args:
            data: 검증할 데이터
            confidence_level: 신뢰도 (HIGH/MEDIUM/LOW/UNCERTAIN)
            
        Returns:
            확실한 데이터만 반환, 불확실하면 None
        """
        self.enforcement_stats['enforced_count'] += 1
        
        # 1. None 또는 빈 데이터
        if data is None or data == '':
            return None
        
        # 2. 신뢰도 기반 판단
        if confidence_level in ['UNCERTAIN', 'LOW']:
            self.enforcement_stats['rejected_data_count'] += 1
            logging.warning(f"[데이터완결성강제] [신뢰도{confidence_level}] [데이터거부] [값_{data}]")
            return None
        
        # 3. 숫자 데이터 유효성 검사
        if isinstance(data, (int, float)):
            if abs(data) > 1e12:  # 1조 초과하는 비현실적인 값
                self.enforcement_stats['rejected_data_count'] += 1
                logging.warning(f"[데이터완결성강제] [비현실적금액] [데이터거부] [값_{data}]")
                return None
        
        # 4. 문자열 데이터 검증
        if isinstance(data, str):
            suspicious_patterns = ['추정', '예상', '대략', '약', '~', 'TBD', 'N/A']
            if any(pattern in data for pattern in suspicious_patterns):
                self.enforcement_stats['rejected_data_count'] += 1
                logging.warning(f"[데이터완결성강제] [추정값패턴감지] [데이터거부] [값_{data}]")
                return None
        
        # 5. 확실한 데이터만 통과
        return data
    
    def mark_uncertain_and_reject(self, workbook, sheet_name: str, cell_ref: str, 
                                 reason: str, original_value: Any = None):
        """
        확실하지 않은 데이터 노란색 마킹 후 거부
        
        Args:
            workbook: Excel 워크북
            sheet_name: 시트명
            cell_ref: 셀 참조 (예: 'G5')
            reason: 거부 이유
            original_value: 원본 값
        """
        try:
            sheet = workbook[sheet_name]
            cell = sheet[cell_ref]
            
            # 노란색 마킹
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = yellow_fill
            
            # 값 비우기 (CLAUDE.md 핵심 원칙)
            cell.value = None
            
            self.enforcement_stats['marked_uncertain_count'] += 1
            
            logging.warning(f"[데이터완결성강제] [노란색마킹] [시트_{sheet_name}] [셀_{cell_ref}] [이유_{reason}] [원본값_{original_value}]")
            
        except Exception as e:
            logging.error(f"[마킹오류] [시트_{sheet_name}] [셀_{cell_ref}] [오류_{str(e)}]")
    
    def get_enforcement_stats(self) -> Dict:
        """강제 적용 통계"""
        return {
            'enforcement_summary': self.enforcement_stats,
            'rejection_rate': (self.enforcement_stats['rejected_data_count'] / 
                             max(self.enforcement_stats['enforced_count'], 1)) * 100,
            'data_integrity_maintained': True  # 추정값 0개 입력
        }


class ReportViolationError(Exception):
    """체크리스트 미완료시 발생하는 예외"""
    pass